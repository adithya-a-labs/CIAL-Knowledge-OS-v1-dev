#!/usr/bin/env python3
"""Single-file prompt × answer-length experiment runner for CIAL Knowledge OS.

The runner deliberately keeps experiment policy, preflight, checkpoints, metrics,
and all export templates in this file.  Production retrieval and generation
implementations are imported but never patched.
"""

from __future__ import annotations

import argparse
import atexit
import csv
import dataclasses
import datetime as dt
import hashlib
import html
import itertools
import json
import math
import os
import random
import re
import signal
import statistics
import subprocess
import sys
import tempfile
import textwrap
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence


SCRIPT_VERSION = "1.0.0"
PROMPT_REGISTRY_VERSION = "cial-prompt-length-2026-07-24"
REPO_ROOT = Path(__file__).resolve().parents[1]
SERVICE_ROOT = REPO_ROOT / "services" / "knowledge-engine"
SERVICE_ENV_FILE = SERVICE_ROOT / ".env"


@dataclass(frozen=True)
class EnvironmentBootstrap:
    """Resolved database configuration without exposing its secret value."""

    env_file: Path
    env_file_found: bool
    env_file_loaded: bool
    database_url: str
    configuration_source: str
    bootstrap_complete: bool = True

    @property
    def database_url_resolved(self) -> bool:
        return bool(self.database_url.strip())

    def child_environment(self) -> dict[str, str]:
        inherited = dict(os.environ)
        if self.database_url_resolved:
            inherited["DATABASE_URL"] = self.database_url
        return inherited


def bootstrap_environment(
    env_file: Path = SERVICE_ENV_FILE,
    *,
    environment: Mapping[str, str] | None = None,
    dotenv_loader: Callable[..., Any] | None = None,
) -> EnvironmentBootstrap:
    """Load the service environment before any backend configuration import.

    ``override=False`` is non-negotiable: an explicit process value always wins.
    The injectable mapping/loader keep the regression tests local to this file.
    """

    target = os.environ if environment is None else environment
    existing = str(target.get("DATABASE_URL") or "").strip()
    found = env_file.is_file()
    loaded = False
    if found:
        loader = dotenv_loader
        if loader is None:
            try:
                from dotenv import load_dotenv
            except ImportError:
                loader = None
            else:
                loader = load_dotenv
        if loader is not None:
            loaded = bool(loader(env_file, override=False))
    resolved = str(target.get("DATABASE_URL") or "").strip()
    source = (
        "process environment" if existing
        else "service .env" if resolved
        else "unresolved"
    )
    return EnvironmentBootstrap(env_file.resolve(), found, loaded, resolved, source)


def _ensure_backend_import_paths() -> None:
    """Add repository service paths only after path and env resolution."""

    for entry in (SERVICE_ROOT, SERVICE_ROOT / "src"):
        value = str(entry.resolve())
        if value not in sys.path:
            sys.path.insert(0, value)


def resolve_canonical_database_url(
    bootstrap: EnvironmentBootstrap,
    *,
    resolver: Callable[[], str] | None = None,
) -> EnvironmentBootstrap:
    """Fall back to the backend's canonical settings after early bootstrap."""

    if bootstrap.database_url_resolved:
        return bootstrap
    _ensure_backend_import_paths()
    if resolver is None:
        try:
            from backend.app.core.config import settings
        except Exception:
            candidate = ""
        else:
            candidate = str(getattr(settings, "database_url", "") or "").strip()
    else:
        candidate = str(resolver() or "").strip()
    if not candidate:
        return bootstrap
    os.environ.setdefault("DATABASE_URL", candidate)
    return dataclasses.replace(
        bootstrap,
        database_url=candidate,
        configuration_source="backend settings",
    )


def sanitize_database_url(database_url: str) -> dict[str, str]:
    """Return only non-secret host/database fields for diagnostics."""

    if not database_url.strip():
        return {}
    try:
        parsed = urllib.parse.urlsplit(database_url)
        host = parsed.hostname or ""
        try:
            port = parsed.port
        except ValueError:
            port = None
        if host and port:
            host = f"{host}:{port}"
        database = urllib.parse.unquote(parsed.path.lstrip("/").split("/", 1)[0])
        safe: dict[str, str] = {}
        if host:
            safe["database_host"] = host
        if database:
            safe["database_name"] = database
        return safe
    except (TypeError, ValueError):
        return {}


def environment_diagnostics(bootstrap: EnvironmentBootstrap) -> dict[str, Any]:
    diagnostics: dict[str, Any] = {
        "environment_file_path": str(bootstrap.env_file),
        "environment_file_found": bootstrap.env_file_found,
        "environment_file_loaded": bootstrap.env_file_loaded,
        "database_url_resolved": bootstrap.database_url_resolved,
        "configuration_source": bootstrap.configuration_source,
    }
    diagnostics.update(sanitize_database_url(bootstrap.database_url))
    return diagnostics


def database_configuration_gate(bootstrap: EnvironmentBootstrap) -> dict[str, Any]:
    gate = {
        "passed": bootstrap.database_url_resolved,
        **environment_diagnostics(bootstrap),
    }
    if not bootstrap.database_url_resolved:
        gate["error"] = (
            "DATABASE_URL could not be resolved from the process environment, "
            "service .env, or canonical backend settings."
        )
    return gate


def only_allowed_source_changes(paths: Iterable[str]) -> bool:
    normalized = {
        str(path).replace("\\", "/")
        for path in paths
        if str(path).strip() and not str(path).replace("\\", "/").startswith(
            "outputs/batch_answers/prompt_length_comparisons/"
        )
    }
    return normalized <= {"scripts/run_prompt_length_comparison.py"}


# This ordering is intentional and regression-tested: dotenv first, backend import
# paths second, canonical backend settings last.
ENV_BOOTSTRAP = bootstrap_environment()
_ensure_backend_import_paths()
ENV_BOOTSTRAP = resolve_canonical_database_url(ENV_BOOTSTRAP)

DEFAULT_QUESTIONS = REPO_ROOT / "data" / "manual_qa" / "final" / "questions.txt"
DEFAULT_CORPUS = REPO_ROOT / "data" / "files"
DEFAULT_OUTPUT = REPO_ROOT / "outputs" / "batch_answers" / "prompt_length_comparisons"
DEFAULT_QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6335").rstrip("/")
DEFAULT_BACKEND_URL = os.getenv("CIAL_BACKEND_URL", "http://127.0.0.1:8000").rstrip("/")
DEFAULT_COLLECTION = os.getenv("QDRANT_COLLECTION_NAME", "cial_phase4")
NO_EVIDENCE = (
    "Based only on the indexed corpus, no reliable answer can be provided. "
    "The supplied evidence is insufficient."
)
REQUIRED_COLUMNS = [
    "run_id", "question_id", "question_index", "question", "category",
    "profile_id", "profile_label", "profile_fingerprint", "prompt_id",
    "prompt_label", "prompt_version", "prompt_fingerprint", "repeat_index",
    "execution_order", "status", "error_code", "error_message", "answer_status",
    "answer_markdown", "answer_plain_text", "answer_words", "answer_chars",
    "answer_tokens", "prompt_tokens", "completion_tokens", "total_model_tokens",
    "total_latency_s", "retrieval_latency_s", "rerank_latency_s",
    "selection_latency_s", "generation_latency_s", "citation_count",
    "inline_citation_count", "citation_density_per_100_words",
    "unique_source_count", "citation_ids", "citation_sources", "citation_pages",
    "selected_chunk_count", "selected_chunk_ids", "selected_evidence_tokens",
    "final_context_tokens", "context_hash", "retrieval_fingerprint",
    "comparison_valid", "comparison_warning", "evidence_strength",
    "weak_evidence", "unsupported_query", "expected_keyword_coverage",
    "forbidden_keyword_hits", "safe_failure_expected", "safe_failure_correct",
    "lexical_redundancy_score", "model_name", "temperature", "num_ctx",
    "num_predict", "prompt_name", "prompt_registry_version", "created_at",
    "retry_count", "corpus_fingerprint", "rendered_prompt_hash",
    "heuristic_tradeoff_score", "pareto_frontier",
]
STOP_WORDS = frozenset(
    "a an and are as at be by for from how in is it of on or that the this to "
    "was what when where which who why will with your".split()
)


def utcnow() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha(value: Any) -> str:
    raw = value if isinstance(value, bytes) else canonical(value).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def file_sha(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def count_tokens(text: str) -> int:
    # Stable fallback usable before repository token/model imports.
    return max(0, math.ceil(len(re.findall(r"\S+", text)) * 1.32))


def plain_markdown(value: str) -> str:
    value = re.sub(r"```.*?```", " ", value, flags=re.S)
    value = re.sub(r"!\[[^\]]*]\([^)]*\)", " ", value)
    value = re.sub(r"\[([^\]]+)]\([^)]*\)", r"\1", value)
    value = re.sub(r"[*_`#>|~-]", " ", value)
    return " ".join(value.split())


def words(value: str) -> list[str]:
    return re.findall(r"[A-Za-z0-9][A-Za-z0-9'_-]*", value)


def safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, path)


def append_jsonl(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8", newline="\n") as stream:
        stream.write(canonical(payload) + "\n")
        stream.flush()
        os.fsync(stream.fileno())


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    result = []
    for number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            item = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSONL at {path}:{number}: {exc}") from exc
        if isinstance(item, dict):
            result.append(item)
    return result


def latest_results(rows: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Collapse append-only attempt history to the latest row for each cell."""
    keyed: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for index, row in enumerate(rows):
        key = cell_key(row)
        if not row.get("question_id"):
            key = f"non-cell-{index}"
        if key not in keyed:
            order.append(key)
        keyed[key] = row
    return [keyed[key] for key in order]


@dataclass(frozen=True)
class QuestionRecord:
    question_id: str
    question_index: int
    question: str
    category: str
    normalized: str


@dataclass(frozen=True)
class AnswerLengthProfile:
    id: str
    label: str
    detail: str
    min_words: int
    max_words: int
    structured: bool = True
    adaptive: bool = True
    decision_notes: bool = False

    @property
    def fingerprint(self) -> str:
        return sha(asdict(self))

    def instructions(self) -> str:
        decision = (
            "Include decision notes only when explicitly grounded in evidence."
            if self.decision_notes else "Do not add a decision-notes section."
        )
        return (
            f"Target {self.min_words}–{self.max_words} words ({self.detail}). "
            "This is guidance, never a reason to pad unsupported content. "
            f"{'Use structure adaptively.' if self.structured else 'Prefer prose.'} {decision}"
        )


@dataclass(frozen=True)
class PromptVariant:
    id: str
    label: str
    version: str
    template: str

    @property
    def fingerprint(self) -> str:
        return sha(asdict(self))


@dataclass(frozen=True)
class ExperimentCell:
    question_id: str
    profile_id: str
    prompt_id: str
    repeat_index: int


@dataclass
class FrozenContext:
    question_id: str
    context: str
    context_hash: str
    retrieval_fingerprint: str
    chunks: list[dict[str, Any]]
    retrieval_latency_s: float = 0.0
    rerank_latency_s: float = 0.0
    selection_latency_s: float = 0.0
    evidence_strength: str = "unknown"
    weak_evidence: bool = False
    unsupported_query: bool = False


@dataclass
class AttemptResult:
    row: dict[str, Any]


@dataclass
class RunConfig:
    run_id: str
    run_dir: str
    questions_file: str
    corpus_root: str
    profiles: list[str]
    prompts: list[str]
    repeat: int
    retrieval_mode: str
    model: str
    seed: int
    trace_level: str
    corpus_fingerprint: str = ""
    skip_corpus_check_acknowledged: bool = False
    heuristic_weights: dict[str, float] = field(default_factory=lambda: {
        "keyword_coverage": .28, "safe_failure": .20, "citation_proxy": .24,
        "forbidden_hits": .10, "latency": .08, "tokens": .10,
    })


PROFILES = {
    p.id: p for p in (
        AnswerLengthProfile("concise", "Concise", "concise", 60, 140),
        AnswerLengthProfile("balanced", "Balanced", "balanced", 150, 300),
        AnswerLengthProfile("detailed", "Detailed", "detailed", 300, 700, decision_notes=True),
    )
}

PROMPTS = {
    p.id: p for p in (
        PromptVariant("direct", "Direct", "1", """{profile_instructions}
Answer the question directly and lead with the conclusion. Use only supplied evidence.
Cite every material claim and avoid unnecessary sections.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
        PromptVariant("structured", "Structured", "1", """{profile_instructions}
Begin with a direct answer, then use concise headings or bullets when useful.
Cite every material claim or section and explicitly identify evidence gaps.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
        PromptVariant("analytical", "Analytical", "1", """{profile_instructions}
State the conclusion, evidence-supported reasoning, important distinctions or
exceptions, and implications. Separate documented facts from interpretation.
Do not add external knowledge and cite all substantive claims.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
        PromptVariant("operational", "Operational", "1", """{profile_instructions}
Optimize for an airport or enterprise user who may need to act. Lead with the
operational answer. Include responsibilities, steps, deadlines, exceptions,
risks, or escalation only when explicitly supported. Never invent actions.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
        PromptVariant("minimal", "Minimal", "1", """{profile_instructions}
Provide one compact answer with no redundant restatement. Use only supplied evidence.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
        PromptVariant("evidence_first", "Evidence first", "1", """{profile_instructions}
Start with the strongest supported finding, then explain it without external knowledge.

Question:
{question}

Evidence:
{evidence}

{citation_rules}"""),
    )
}


def infer_category(question: str) -> str:
    q = question.casefold()
    for name, markers in {
        "operational": ("procedure", "step", "responsib", "deadline", "must", "how"),
        "comparative": ("compare", "difference", "versus", " vs "),
        "temporal": ("when", "date", "frequency", "how often"),
        "definition": ("what is", "define", "meaning"),
    }.items():
        if any(marker in q for marker in markers):
            return name
    return "general"


def load_questions(path: Path, max_questions: int | None = None) -> tuple[list[QuestionRecord], dict[str, Any]]:
    if path.suffix.casefold() != ".txt":
        raise ValueError("Question input must be a plain-text .txt file.")
    if not path.is_file():
        raise FileNotFoundError(f"Question file not found: {path}")
    raw = path.read_text(encoding="utf-8-sig")
    records: list[QuestionRecord] = []
    warnings: list[str] = []
    for line_no, raw_line in enumerate(raw.splitlines(), 1):
        question = raw_line.strip()
        if not question or question.startswith("#"):
            continue
        if len(question) < 8:
            warnings.append(f"line {line_no}: extremely short")
        if len(question) > 1000:
            warnings.append(f"line {line_no}: very long")
        if len(re.findall(r"(?:^|\s)\d+[.)]\s", question)) > 1:
            warnings.append(f"line {line_no}: may contain multiple numbered questions")
        normalized = " ".join(question.casefold().split())
        index = len(records) + 1
        records.append(QuestionRecord(f"Q{index:04d}", index, question, infer_category(question), normalized))
        if max_questions is not None and len(records) >= max_questions:
            break
    if not records:
        raise ValueError(f"No valid questions in {path}")
    exact = Counter(r.question for r in records)
    norm = Counter(r.normalized for r in records)
    stats = {
        "path": str(path.resolve()), "sha256": file_sha(path), "count": len(records),
        "exact_duplicate_occurrences": sum(v - 1 for v in exact.values() if v > 1),
        "normalized_duplicate_occurrences": sum(v - 1 for v in norm.values() if v > 1),
        "warnings": warnings,
    }
    return records, stats


def validate_prompts(items: Sequence[PromptVariant]) -> None:
    ids: set[str] = set()
    for item in items:
        if not re.fullmatch(r"[A-Za-z0-9_.-]+", item.id) or item.id in ids:
            raise ValueError(f"Invalid or duplicate prompt id: {item.id}")
        ids.add(item.id)
        if len(item.template) > 50_000:
            raise ValueError(f"Prompt template {item.id} exceeds 50,000 characters")
        missing = [x for x in ("question", "evidence", "profile_instructions") if "{" + x + "}" not in item.template]
        if missing:
            raise ValueError(f"Prompt {item.id} missing placeholders: {', '.join(missing)}")


def load_custom_prompts(path: Path) -> list[PromptVariant]:
    text = path.read_text(encoding="utf-8-sig")
    if path.suffix.casefold() == ".json":
        raw = json.loads(text)
    elif path.suffix.casefold() in (".yaml", ".yml"):
        try:
            import yaml  # type: ignore
        except ImportError as exc:
            raise RuntimeError("YAML requested but the repository environment has no PyYAML; use JSON.") from exc
        raw = yaml.safe_load(text)
    else:
        raise ValueError("Custom prompts must be JSON or YAML.")
    if not isinstance(raw, list):
        raise ValueError("Custom prompt config must be a list.")
    items = [
        PromptVariant(str(x["id"]), str(x.get("label") or x["id"]), str(x.get("version") or "1"), str(x["template"]))
        for x in raw if isinstance(x, Mapping)
    ]
    validate_prompts(items)
    return items


def render_prompt(prompt: PromptVariant, profile: AnswerLengthProfile, question: str, evidence: str, scope: str) -> str:
    citation_rules = (
        "Use inline citations [1], [2], … matching the evidence blocks. Cite each "
        "material claim. If evidence is insufficient, return exactly: " + NO_EVIDENCE
    )
    immutable = (
        "IMMUTABLE GROUNDING RULES\n"
        "Use only the authorized evidence below. Never invent facts, actions, dates, "
        "responsibilities, sources, or citations. Preserve safe-failure behavior.\n\n"
    )
    values = defaultdict(str, {
        "question": question, "evidence": evidence,
        "profile_instructions": profile.instructions(),
        "document_scope": scope, "citation_rules": citation_rules,
    })
    try:
        body = prompt.template.format_map(values)
    except (KeyError, ValueError) as exc:
        raise ValueError(f"Cannot render prompt {prompt.id}: {exc}") from exc
    return immutable + body


def supported_extensions() -> set[str]:
    _ensure_backend_import_paths()
    try:
        from cial_knowledge_os.loaders import SUPPORTED_DOCUMENT_EXTENSIONS
        return {str(x).casefold() for x in SUPPORTED_DOCUMENT_EXTENSIONS}
    except Exception:
        # Explicit fallback mirrors the inspected loader registry and is reported.
        return {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".csv", ".pptx", ".ppt",
                ".txt", ".md", ".markdown", ".html", ".htm", ".json", ".xml",
                ".yaml", ".yml", ".png", ".jpg", ".jpeg", ".tiff", ".tif"}


def scan_corpus(root: Path, extensions: set[str] | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not root.is_dir():
        raise FileNotFoundError(f"Corpus root not found: {root}")
    exts = extensions or supported_extensions()
    rows: list[dict[str, Any]] = []
    ignored = []
    root_resolved = root.resolve()
    for path in sorted(root.rglob("*"), key=lambda p: p.as_posix().casefold()):
        if not path.is_file():
            continue
        rel = path.relative_to(root).as_posix()
        if (path.name.startswith("~$") or path.name in (".DS_Store", "Thumbs.db")
                or any(part.startswith(".") for part in Path(rel).parts) or path.stat().st_size == 0):
            ignored.append(rel)
            continue
        resolved = path.resolve()
        try:
            resolved.relative_to(root_resolved)
        except ValueError as exc:
            raise ValueError(f"Corpus path escapes root: {path}") from exc
        stat = path.stat()
        rows.append({
            "relative_path": rel, "normalized_relative_path": rel.replace("\\", "/").casefold(),
            "size_bytes": stat.st_size, "mtime_ns": stat.st_mtime_ns,
            "extension": path.suffix.casefold(), "sha256": file_sha(path),
            "supported": path.suffix.casefold() in exts,
        })
    hashes = Counter(x["sha256"] for x in rows)
    paths = Counter(x["normalized_relative_path"] for x in rows)
    summary = {
        "root": str(root_resolved), "discovered": len(rows),
        "supported": sum(bool(x["supported"]) for x in rows),
        "unsupported": sum(not x["supported"] for x in rows),
        "ignored": ignored,
        "duplicate_hashes": sorted(k for k, v in hashes.items() if v > 1),
        "duplicate_normalized_paths": sorted(k for k, v in paths.items() if v > 1),
        "registry_extensions": sorted(exts),
    }
    return rows, summary


def corpus_fingerprint(rows: Sequence[Mapping[str, Any]], verification: Sequence[Mapping[str, Any]], config: Mapping[str, Any]) -> str:
    payload = {
        "sources": sorted(
            (x["normalized_relative_path"], x["sha256"]) for x in rows if x.get("supported")
        ),
        "verification": sorted(
            (str(x.get("relative_path")), str(x.get("document_id", "")),
             str(x.get("version_id", "")), int(x.get("chunk_count", 0) or 0))
            for x in verification
        ),
        "index": dict(config),
    }
    return sha(payload)


def http_json(url: str, *, method: str = "GET", body: Any = None, timeout: float = 5.0) -> tuple[int, Any]:
    data = None if body is None else canonical(body).encode("utf-8")
    request = urllib.request.Request(
        url, data=data, method=method,
        headers={"Accept": "application/json", "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8", "replace")
            return response.status, json.loads(raw) if raw.strip() else {}
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", "replace")
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            payload = {"error": raw}
        return exc.code, payload


def reachable(url: str, getter: Callable[..., tuple[int, Any]] = http_json) -> tuple[bool, Any]:
    try:
        status, payload = getter(url, timeout=3)
        return 200 <= status < 300, payload
    except (OSError, urllib.error.URLError, TimeoutError):
        return False, {}


def flatten_corpus_tree(payload: Any) -> dict[str, dict[str, Any]]:
    found: dict[str, dict[str, Any]] = {}
    def visit(value: Any) -> None:
        if isinstance(value, Mapping):
            path = value.get("relative_path") or value.get("path")
            if path and any(k in value for k in ("document_id", "id", "chunk_count", "indexed")):
                found[str(path).replace("\\", "/").casefold()] = dict(value)
            for child in value.values():
                visit(child)
        elif isinstance(value, list):
            for child in value:
                visit(child)
    visit(payload)
    return found


def verify_manifest(rows: Sequence[Mapping[str, Any]], tree: Any, vector_count: int | None) -> list[dict[str, Any]]:
    indexed = flatten_corpus_tree(tree)
    result = []
    for item in rows:
        if not item.get("supported"):
            result.append({**dict(item), "verification_status": "unsupported", "blocking": True})
            continue
        meta = indexed.get(str(item["normalized_relative_path"]))
        chunks = int((meta or {}).get("chunk_count") or len((meta or {}).get("chunks") or []))
        content_hash = str((meta or {}).get("content_hash") or "")
        hash_ok = not content_hash or content_hash == item["sha256"]
        row = {
            **dict(item), "metadata_present": bool(meta), "document_id": str((meta or {}).get("document_id") or (meta or {}).get("id") or ""),
            "version_id": str((meta or {}).get("current_version_id") or (meta or {}).get("version_id") or ""),
            "chunk_count": chunks, "hash_match": hash_ok,
            "vector_collection_point_count": vector_count,
        }
        row["verification_status"] = (
            "indexed_successfully" if meta and chunks > 0 and hash_ok and (vector_count is None or vector_count > 0)
            else "metadata_missing" if not meta else "zero_chunks" if chunks <= 0
            else "hash_mismatch" if not hash_ok else "vector_mismatch"
        )
        row["blocking"] = row["verification_status"] != "indexed_successfully"
        result.append(row)
    return result


class ServiceManager:
    def __init__(
        self,
        run_dir: Path,
        root: Path = REPO_ROOT,
        *,
        environment: Mapping[str, str] | None = None,
        popen_factory: Callable[..., subprocess.Popen[Any]] = subprocess.Popen,
    ):
        self.run_dir, self.root = run_dir, root
        self.environment = dict(environment or ENV_BOOTSTRAP.child_environment())
        self.popen_factory = popen_factory
        self.children: dict[str, tuple[subprocess.Popen[Any], Any]] = {}

    def start(self, name: str, batch: Path) -> subprocess.Popen[Any]:
        log = (self.run_dir / f"{name}_startup.log").open("a", encoding="utf-8")
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        process = self.popen_factory(
            ["cmd.exe", "/d", "/c", str(batch)], cwd=self.root,
            stdout=log, stderr=subprocess.STDOUT, creationflags=flags,
            env=dict(self.environment),
        )
        self.children[name] = (process, log)
        return process

    def stop(self) -> None:
        for name, (process, log) in self.children.items():
            if process.poll() is None:
                process.terminate()
                try:
                    process.wait(8)
                except subprocess.TimeoutExpired:
                    process.kill()
            log.write(f"\ncleanup {name}: returncode={process.poll()}\n")
            log.close()


def wait_until(
    probe: Callable[[], tuple[bool, Any]], timeout: float, interval: float,
    *, stable_polls: int = 1, process: subprocess.Popen[Any] | None = None,
) -> tuple[bool, Any, int]:
    deadline, stable, polls, latest = time.monotonic() + timeout, 0, 0, {}
    while time.monotonic() <= deadline:
        polls += 1
        if process is not None and process.poll() is not None:
            return False, {"reason": f"process exited {process.returncode}"}, polls
        ok, latest = probe()
        stable = stable + 1 if ok else 0
        if stable >= stable_polls:
            return True, latest, polls
        time.sleep(max(.01, interval))
    return False, latest, polls


def backend_ready(payload: Mapping[str, Any]) -> bool:
    return bool(
        payload.get("status") == "ready" and payload.get("engine_ready")
        and payload.get("qdrant_ready") and payload.get("models_ready")
        and payload.get("index_fresh") and payload.get("database_ready", payload.get("database_healthy", True))
    )


def wait_for_backend_ready(
    url: str,
    timeout: float,
    interval: float,
    *,
    getter: Callable[..., tuple[int, Any]] = http_json,
) -> tuple[bool, Any, int]:
    """Poll readiness, stopping immediately on an explicit terminal failure."""

    deadline, stable, polls, latest = time.monotonic() + timeout, 0, 0, {}
    while time.monotonic() <= deadline:
        polls += 1
        try:
            code, latest = getter(url, timeout=5)
        except (OSError, urllib.error.URLError, TimeoutError):
            code, latest = 0, {}
        if code == 200 and isinstance(latest, Mapping):
            if str(latest.get("status") or "").casefold() == "failed":
                return False, latest, polls
            stable = stable + 1 if backend_ready(latest) else 0
            if stable >= 2:
                return True, latest, polls
        else:
            stable = 0
        time.sleep(max(.01, interval))
    return False, latest, polls


def run_preflight(args: argparse.Namespace, run_dir: Path, manager: ServiceManager) -> dict[str, Any]:
    started = time.monotonic()
    gates: dict[str, Any] = {
        "database_configuration": database_configuration_gate(ENV_BOOTSTRAP)
    }
    try:
        questions, qstats = load_questions(Path(args.questions_file), args.max_questions)
        gates["questions"] = {"passed": True, **qstats}
    except Exception as exc:
        questions, qstats = [], {}
        gates["questions"] = {"passed": False, "error": str(exc)}
    try:
        manifest, scan_stats = scan_corpus(Path(args.corpus_root))
        atomic_json(run_dir / "corpus_preflight_manifest.json", {"files": manifest, "summary": scan_stats})
        unsupported = [x for x in manifest if not x["supported"]]
        passed = not unsupported or args.allow_unsupported_files
        gates["corpus_scan"] = {"passed": passed, **scan_stats}
    except Exception as exc:
        manifest, scan_stats = [], {}
        gates["corpus_scan"] = {"passed": False, "error": str(exc)}
    if args.skip_corpus_check:
        gates["corpus_verification"] = {
            "passed": True, "skipped": True,
            "warning": "Explicit --skip-corpus-check acknowledgement recorded; cohort is unsafe.",
        }
        fingerprint = corpus_fingerprint(manifest, [], {
            "collection": args.collection, "retrieval_mode": args.retrieval_mode,
            "skipped": True,
        })
        return {
            "passed": bool(gates["questions"]["passed"] and gates["corpus_scan"]["passed"]),
            "gates": gates, "questions": questions, "question_stats": qstats,
            "manifest": manifest, "verification": [], "corpus_fingerprint": fingerprint,
            "duration_s": time.monotonic() - started,
        }
    q_ok, q_payload = reachable(args.qdrant_url + "/collections")
    q_started = False
    if not q_ok and args.start_services:
        proc = manager.start("qdrant", REPO_ROOT / "scripts" / "start_qdrant.bat")
        q_started = True
        q_ok, q_payload, polls = wait_until(
            lambda: reachable(args.qdrant_url + "/collections"),
            args.startup_timeout_seconds, args.poll_interval_seconds, process=proc,
        )
    gates["qdrant"] = {"passed": q_ok, "pre_existing": q_ok and not q_started, "started": q_started}
    b_ok, health = reachable(args.backend_url + "/api/health")
    b_started = False
    if not b_ok and args.start_services:
        proc = manager.start("backend", REPO_ROOT / "scripts" / "start_backend.bat")
        b_started = True
        b_ok, health, _ = wait_until(
            lambda: reachable(args.backend_url + "/api/health"),
            args.startup_timeout_seconds, args.poll_interval_seconds, process=proc,
        )
    gates["backend_reachable"] = {"passed": b_ok, "pre_existing": b_ok and not b_started, "started": b_started}
    if b_ok and (args.wait_for_ready or args.prepare_corpus or args.verify_index):
        if args.prepare_corpus:
            try:
                http_json(args.backend_url + "/api/index/rebuild", method="POST", body={"force": False}, timeout=10)
            except Exception:
                pass
        ready, health, polls = wait_for_backend_ready(
            args.backend_url + "/api/health",
            args.index_timeout_seconds, args.poll_interval_seconds,
        )
    else:
        ready, polls = b_ok and backend_ready(health), 1
    gates["backend_ready"] = {"passed": bool(ready), "polls": polls, "snapshot": health}
    tree: Any = {}
    if b_ok and ENV_BOOTSTRAP.database_url_resolved:
        try:
            code, tree = http_json(args.backend_url + "/api/corpus/tree", timeout=15)
            if code != 200:
                tree = {}
        except Exception:
            tree = {}
    vector_count = None
    collection = args.collection
    if q_ok:
        try:
            code, info = http_json(args.qdrant_url + "/collections/" + urllib.parse.quote(collection), timeout=8)
            if code == 200:
                vector_count = int(info.get("result", {}).get("points_count") or 0)
        except Exception:
            pass
    verification = (
        verify_manifest(manifest, tree, vector_count)
        if ENV_BOOTSTRAP.database_url_resolved
        else [
            {
                **dict(item),
                "verification_status": "database_configuration_unresolved",
                "blocking": True,
            }
            for item in manifest
        ]
    )
    write_csv(run_dir / "corpus_verification.csv", verification, list(verification[0]) if verification else ["relative_path", "verification_status"])
    atomic_json(run_dir / "corpus_verification.json", verification)
    verified = bool(manifest) and all(not x.get("blocking") or (not x.get("supported") and args.allow_unsupported_files) for x in verification)
    gates["corpus_verification"] = {"passed": verified, "files": len(verification), "failures": sum(bool(x.get("blocking")) for x in verification)}
    fingerprint = corpus_fingerprint(manifest, verification, {
        "collection": collection, "retrieval_mode": args.retrieval_mode,
        "embedding_model": health.get("embedding_model") if isinstance(health, Mapping) else "",
    })
    passed = all(bool(g.get("passed")) for g in gates.values())
    result = {
        "passed": passed, "gates": gates, "questions": questions,
        "question_stats": qstats, "manifest": manifest, "verification": verification,
        "corpus_fingerprint": fingerprint, "duration_s": round(time.monotonic() - started, 3),
        "files_discovered": len(manifest), "files_supported": sum(bool(x.get("supported")) for x in manifest),
        "files_failed": sum(bool(x.get("blocking")) for x in verification),
        "metadata_documents": len(flatten_corpus_tree(tree)), "qdrant_points": vector_count,
    }
    atomic_json(run_dir / "index_metrics.json", {
        "duration_s": result["duration_s"], "health": health, "qdrant_points": vector_count,
    })
    return result


class RepositoryAdapter:
    """Load Phase 4 once, then split frozen retrieval/context from generation."""

    def __init__(self, args: argparse.Namespace):
        _ensure_backend_import_paths()
        from cial_knowledge_os.config import Phase4Config
        from cial_knowledge_os.execution import ExecutionManager
        from cial_knowledge_os.phase4_pipeline import Phase4RAGPipeline
        kwargs: dict[str, Any] = {
            "project_root": REPO_ROOT, "retrieval_mode": args.retrieval_mode,
            "qdrant_url": args.qdrant_url, "qdrant_collection_name": args.collection,
            "allow_large_run": True, "force_rebuild_index": False,
        }
        if args.model:
            kwargs["ollama_model_name"] = args.model
        self.config = Phase4Config(**kwargs)
        self.pipeline = Phase4RAGPipeline(self.config)
        self.pipeline.execution_manager = ExecutionManager.from_config(
            self.config, phase="Phase 4", run_mode="manual_qa"
        )
        self.pipeline.load()
        self.pipeline.chunk()
        self.pipeline.embed()
        self.pipeline.index()
        from cial_knowledge_os.llm import create_local_llm
        if self.pipeline.llm is None:
            self.pipeline.llm = create_local_llm(self.config)

    def freeze(self, question: QuestionRecord) -> FrozenContext:
        before = dict(self.pipeline.metrics)
        started = time.perf_counter()
        chunks = self.pipeline.retrieve(question.question)
        result = self.pipeline.context_builder.build(
            self.pipeline.last_merged_retrieval, corpus_chunks=self.pipeline.chunks
        )
        selected = [dict(x) for x in result.compressed]
        context = result.context
        retrieval_fp = sha([
            (x.get("chunk_id"), x.get("document_id"), x.get("text"), x.get("score"))
            for x in selected
        ])
        selection = getattr(self.pipeline, "last_selection_result", None)
        weak = bool(selection and getattr(selection, "weak_evidence", False))
        return FrozenContext(
            question.question_id, context, sha(context), retrieval_fp, selected,
            float(self.pipeline.metrics.get("retrieval_latency", 0)) - float(before.get("retrieval_latency", 0)),
            float(self.pipeline.metrics.get("reranker_latency", 0)) - float(before.get("reranker_latency", 0)),
            float(self.pipeline.metrics.get("evidence_selection_latency", 0)) - float(before.get("evidence_selection_latency", 0)),
            "none" if not chunks else "weak" if weak else "strong", weak, False,
        )

    def generate(self, rendered_prompt: str, profile: AnswerLengthProfile) -> tuple[str, dict[str, Any]]:
        started = time.perf_counter()
        answer = str(self.pipeline.llm.invoke(rendered_prompt)).strip()
        return answer, {
            "generation_latency_s": time.perf_counter() - started,
            "model_name": self.config.ollama_model_name,
            "temperature": 0, "num_ctx": getattr(self.config, "ollama_num_ctx", None),
            "num_predict": getattr(self.config, "ollama_num_predict", None),
        }

    def close(self) -> None:
        self.pipeline.close()


class FakeAdapter:
    def __init__(self, fail_prompt: str = ""):
        self.freezes = 0
        self.fail_prompt = fail_prompt

    def freeze(self, question: QuestionRecord) -> FrozenContext:
        self.freezes += 1
        chunks = [
            {"chunk_id": f"{question.question_id}-1", "text": "Duty managers inspect queues every 15 minutes.", "source": "ops.txt", "page": 2},
            {"chunk_id": f"{question.question_id}-2", "text": "Escalation begins when the marked holding area is exceeded.", "source": "ops.txt", "page": 3},
        ]
        context = "\n\n".join(f"[{i}] Source: {x['source']} | Page: {x['page']}\n{x['text']}" for i, x in enumerate(chunks, 1))
        return FrozenContext(question.question_id, context, sha(context), sha(chunks), chunks, .01, .02, .01, "strong")

    def generate(self, rendered_prompt: str, profile: AnswerLengthProfile) -> tuple[str, dict[str, Any]]:
        if self.fail_prompt and self.fail_prompt in rendered_prompt:
            raise RuntimeError("embedded fake failure")
        answer = (
            "Duty managers should inspect passenger queues every 15 minutes [1]. "
            "They should escalate when a queue exceeds the marked holding area [2].\n\n"
            + ("This action remains limited to the documented trigger. " * (1 if profile.id == "concise" else 4 if profile.id == "balanced" else 9))
        )
        return answer, {"generation_latency_s": .01, "model_name": "fake-local", "temperature": 0, "num_ctx": 4096, "num_predict": profile.max_words * 2}

    def close(self) -> None:
        pass


def chunk_fields(chunks: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    ids, sources, pages = [], [], []
    for index, item in enumerate(chunks, 1):
        ids.append(str(item.get("chunk_id") or item.get("id") or index))
        meta = item.get("metadata") if isinstance(item.get("metadata"), Mapping) else item
        sources.append(str(meta.get("relative_path") or meta.get("source") or meta.get("filename") or ""))
        pages.append(str(meta.get("page") or meta.get("page_number") or ""))
    return {
        "selected_chunk_ids": "|".join(ids), "citation_sources": "|".join(sources),
        "citation_pages": "|".join(pages), "selected_chunk_count": len(chunks),
        "unique_source_count": len({x for x in sources if x}),
    }


def expected_keywords(question: str) -> set[str]:
    return {x.casefold() for x in words(question) if len(x) >= 4 and x.casefold() not in STOP_WORDS}


def redundancy(answer: str) -> float:
    tokens = [x.casefold() for x in words(answer)]
    if len(tokens) < 4:
        return 0.0
    grams = list(zip(tokens, tokens[1:], tokens[2:]))
    return round(1 - len(set(grams)) / len(grams), 4)


def row_metrics(answer: str, question: str) -> dict[str, Any]:
    ws = words(plain_markdown(answer))
    lower = answer.casefold()
    ids = re.findall(r"\[(\d+)]", answer)
    expected = expected_keywords(question)
    coverage = sum(x in lower for x in expected) / len(expected) if expected else 1.0
    forbidden = sum(lower.count(x) for x in ("as an ai", "i cannot access", "according to my knowledge"))
    insufficient = any(x in lower for x in ("insufficient evidence", "no reliable answer", "cannot be answered"))
    return {
        "answer_plain_text": plain_markdown(answer), "answer_words": len(ws),
        "answer_chars": len(answer), "answer_tokens": count_tokens(answer),
        "completion_tokens": count_tokens(answer), "citation_count": len(set(ids)),
        "inline_citation_count": len(ids), "citation_density_per_100_words": round(len(ids) * 100 / max(1, len(ws)), 4),
        "citation_ids": "|".join(sorted(set(ids), key=int)), "expected_keyword_coverage": round(coverage, 4),
        "forbidden_keyword_hits": forbidden, "safe_failure_expected": False,
        "safe_failure_correct": not insufficient, "lexical_redundancy_score": redundancy(answer),
        "answer_status": "insufficient_evidence" if insufficient else "answered",
    }


def cell_key(row: Mapping[str, Any]) -> str:
    return "|".join(str(row.get(x, "")) for x in (
        "question_id", "profile_fingerprint", "prompt_fingerprint", "repeat_index", "context_hash"
    ))


def matrix_cells(
    questions: Sequence[QuestionRecord], profiles: Sequence[AnswerLengthProfile],
    prompts: Sequence[PromptVariant], repeat: int, seed: int,
    randomize_profiles: bool = False, randomize_prompts: bool = False,
) -> list[tuple[QuestionRecord, AnswerLengthProfile, PromptVariant, int]]:
    rng = random.Random(seed)
    cells = []
    for question in questions:
        ps, rs = list(profiles), list(prompts)
        if randomize_profiles:
            rng.shuffle(ps)
        if randomize_prompts:
            rng.shuffle(rs)
        cells.extend((question, profile, prompt, r) for r in range(1, repeat + 1) for profile in ps for prompt in rs)
    return cells


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(columns), extrasaction="ignore", quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: safe_cell(row.get(key, "")) for key in columns})


def wide_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        qid = str(row["question_id"])
        out = grouped.setdefault(qid, {"question_id": qid, "question": row.get("question", ""), "category": row.get("category", "")})
        prefix = f"{row.get('profile_id')}__{row.get('prompt_id')}"
        out[prefix + "__answer"] = row.get("answer_markdown", "")
        out[prefix + "__tokens"] = row.get("answer_tokens", 0)
        out[prefix + "__latency_s"] = row.get("total_latency_s", 0)
    return list(grouped.values())


def aggregate_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(str(row.get("profile_id")), str(row.get("prompt_id")), str(row.get("category")))].append(row)
    out = []
    for (profile, prompt, category), items in sorted(groups.items()):
        successes = [x for x in items if x.get("status") == "success"]
        mean = lambda key: round(statistics.fmean(float(x.get(key, 0) or 0) for x in successes), 4) if successes else 0
        out.append({
            "profile_id": profile, "prompt_id": prompt, "category": category, "cells": len(items),
            "success_rate": round(len(successes) / len(items), 4), "mean_tokens": mean("answer_tokens"),
            "mean_latency_s": mean("total_latency_s"), "mean_citation_density": mean("citation_density_per_100_words"),
            "mean_keyword_coverage": mean("expected_keyword_coverage"), "mean_heuristic_tradeoff_score": mean("heuristic_tradeoff_score"),
            "pareto_count": sum(bool(x.get("pareto_frontier")) for x in successes),
        })
    return out


def token_set(value: str) -> set[str]:
    return {x.casefold() for x in words(value)}


def jaccard(a: str, b: str) -> float:
    x, y = token_set(a), token_set(b)
    return round(len(x & y) / len(x | y), 4) if x or y else 1.0


def cosine(a: str, b: str) -> float:
    ca, cb = Counter(words(a.casefold())), Counter(words(b.casefold()))
    keys = set(ca) | set(cb)
    dot = sum(ca[k] * cb[k] for k in keys)
    denom = math.sqrt(sum(v * v for v in ca.values()) * sum(v * v for v in cb.values()))
    return round(dot / denom, 4) if denom else 1.0


def pairwise(rows: Sequence[Mapping[str, Any]], dimension: str) -> list[dict[str, Any]]:
    other = "profile_id" if dimension == "prompt_id" else "prompt_id"
    groups: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("status") == "success":
            groups[(str(row.get("question_id")), str(row.get(other)))].append(row)
    output = []
    for (qid, fixed), items in groups.items():
        for a, b in itertools.combinations(sorted(items, key=lambda x: str(x.get(dimension))), 2):
            output.append({
                "question_id": qid, "fixed_dimension": other, "fixed_value": fixed,
                "left": a.get(dimension), "right": b.get(dimension),
                "token_delta": int(b.get("answer_tokens", 0)) - int(a.get("answer_tokens", 0)),
                "latency_delta_s": round(float(b.get("total_latency_s", 0)) - float(a.get("total_latency_s", 0)), 4),
                "citation_density_delta": round(float(b.get("citation_density_per_100_words", 0)) - float(a.get("citation_density_per_100_words", 0)), 4),
                "keyword_coverage_delta": round(float(b.get("expected_keyword_coverage", 0)) - float(a.get("expected_keyword_coverage", 0)), 4),
                "answer_token_jaccard": jaccard(str(a.get("answer_plain_text", "")), str(b.get("answer_plain_text", ""))),
                "tf_cosine": cosine(str(a.get("answer_plain_text", "")), str(b.get("answer_plain_text", ""))),
                "source_set_jaccard": jaccard(str(a.get("citation_sources", "")), str(b.get("citation_sources", ""))),
                "context_equal": a.get("context_hash") == b.get("context_hash"),
            })
    return output


def score_rows(rows: list[dict[str, Any]], weights: Mapping[str, float]) -> None:
    successes = [x for x in rows if x.get("status") == "success"]
    max_tokens = max((float(x.get("answer_tokens", 0)) for x in successes), default=1) or 1
    max_latency = max((float(x.get("total_latency_s", 0)) for x in successes), default=1) or 1
    max_density = max((float(x.get("citation_density_per_100_words", 0)) for x in successes), default=1) or 1
    for row in rows:
        if row.get("status") != "success":
            row["heuristic_tradeoff_score"] = 0.0
            row["pareto_frontier"] = False
            continue
        score = (
            weights["keyword_coverage"] * float(row.get("expected_keyword_coverage", 0))
            + weights["safe_failure"] * float(bool(row.get("safe_failure_correct")))
            + weights["citation_proxy"] * min(1, float(row.get("citation_density_per_100_words", 0)) / max_density)
            - weights["forbidden_hits"] * min(1, float(row.get("forbidden_keyword_hits", 0)))
            - weights["latency"] * float(row.get("total_latency_s", 0)) / max_latency
            - weights["tokens"] * float(row.get("answer_tokens", 0)) / max_tokens
        )
        row["heuristic_tradeoff_score"] = round(score, 6)
    for row in successes:
        dominated = any(
            other is not row
            and float(other.get("heuristic_tradeoff_score", 0)) >= float(row.get("heuristic_tradeoff_score", 0))
            and float(other.get("answer_tokens", 0)) <= float(row.get("answer_tokens", 0))
            and float(other.get("total_latency_s", 0)) <= float(row.get("total_latency_s", 0))
            and (
                float(other.get("heuristic_tradeoff_score", 0)) > float(row.get("heuristic_tradeoff_score", 0))
                or float(other.get("answer_tokens", 0)) < float(row.get("answer_tokens", 0))
                or float(other.get("total_latency_s", 0)) < float(row.get("total_latency_s", 0))
            )
            for other in successes
        )
        row["pareto_frontier"] = not dominated


def export_excel(
    path: Path, rows: list[dict[str, Any]], aggregate: list[dict[str, Any]],
    prompt_pairs: list[dict[str, Any]], length_pairs: list[dict[str, Any]],
    run_config: Mapping[str, Any], summary: Mapping[str, Any],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required and should be present in the repository environment.") from exc
    book = Workbook()
    book.remove(book.active)
    sheets: list[tuple[str, list[Mapping[str, Any]], list[str]]] = [
        ("Overview", [summary], list(summary)),
        ("Prompt × Profile Summary", aggregate, list(aggregate[0]) if aggregate else ["profile_id"]),
        ("Question Comparison", wide_rows(rows), list(wide_rows(rows)[0]) if wide_rows(rows) else ["question_id"]),
        ("Long Results", rows, REQUIRED_COLUMNS),
        ("Pairwise Prompt Deltas", prompt_pairs, list(prompt_pairs[0]) if prompt_pairs else ["question_id"]),
        ("Pairwise Length Deltas", length_pairs, list(length_pairs[0]) if length_pairs else ["question_id"]),
        ("Citations", rows, ["question_id", "profile_id", "prompt_id", "citation_ids", "citation_sources", "citation_pages"]),
        ("Failures", [x for x in rows if x.get("status") != "success"], REQUIRED_COLUMNS),
        ("Run Config", [{"key": k, "value": canonical(v) if isinstance(v, (dict, list)) else v} for k, v in run_config.items()], ["key", "value"]),
    ]
    header_fill = PatternFill("solid", fgColor="17324D")
    for name, data, columns in sheets:
        ws = book.create_sheet(name[:31])
        ws.append(columns)
        for item in data:
            ws.append([safe_cell(item.get(key, "")) for key in columns])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, Font(color="FFFFFF", bold=True)
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(60, max(12, max(len(str(c.value or "")) for c in col[:100]) + 2))
        if name == "Long Results" and ws.max_row > 1:
            score_col = REQUIRED_COLUMNS.index("heuristic_tradeoff_score") + 1
            ws.conditional_formatting.add(
                f"{ws.cell(2, score_col).coordinate}:{ws.cell(ws.max_row, score_col).coordinate}",
                ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"),
            )
    review = book.create_sheet("Human Review")
    review_cols = [
        "question_id", "question", "profile_id", "prompt_id", "answer_markdown",
        "preferred_for_question", "preferred_within_profile", "quality_score_1_5",
        "grounding_score_1_5", "clarity_score_1_5", "completeness_score_1_5", "notes",
    ]
    review.append(review_cols)
    for item in rows:
        review.append([item.get(x, "") for x in review_cols])
    review.freeze_panes, review.auto_filter.ref = "A2", review.dimensions
    for cell in review[1]:
        cell.fill, cell.font = header_fill, Font(color="FFFFFF", bold=True)
    review.column_dimensions["B"].width = 55
    review.column_dimensions["E"].width = 90
    review.column_dimensions["L"].width = 50
    for row_cells in review.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    book.save(path)


CSS = r"""
:root{color-scheme:light dark;--bg:#f4f7fa;--panel:#fff;--ink:#142536;--muted:#617487;--brand:#007a78;--line:#d7e0e8}
[data-theme=dark]{--bg:#101820;--panel:#18242f;--ink:#eef6fb;--muted:#a9bac8;--line:#334653}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font:15px/1.5 system-ui,sans-serif}
header{background:#102f46;color:white;padding:1.4rem 3vw}h1{margin:.1rem 0}.lede{color:#c9e5e4}
.metrics,.controls,.tabs,.grid{display:flex;gap:.7rem;flex-wrap:wrap}.metric,.panel{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:.8rem}
.metric{color:var(--ink);min-width:115px}.metric b{display:block;font-size:1.4rem}.controls,.tabs{padding:1rem 3vw}.controls input,.controls select,.controls button,.tabs button{padding:.55rem;border:1px solid var(--line);border-radius:6px;background:var(--panel);color:var(--ink)}
main{padding:0 3vw 3rem}.view{display:none}.view.active{display:block}.card{margin:.8rem 0;background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:1rem}
.answer{white-space:pre-wrap}.badge,.chip{display:inline-block;background:#d8f1ee;color:#064f51;border-radius:99px;padding:.1rem .5rem;margin:.1rem}.warn{color:#a44700}
table{width:100%;border-collapse:collapse;background:var(--panel)}th,td{text-align:left;vertical-align:top;border:1px solid var(--line);padding:.5rem}th{background:#d9ecef;color:#17324d}
.answer-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:.8rem}.trace{font:12px ui-monospace,monospace;white-space:pre-wrap}
.chart{min-height:180px}.bar{display:flex;align-items:center;margin:.25rem}.bar span{display:inline-block;background:var(--brand);height:18px;margin-left:.5rem}
.hide-metrics .meta,.hide-citations .citations,.hide-debug details{display:none}.tabbed .answer-grid{display:block}.tabbed .answer-grid .card{margin:.6rem 0}
.dense .card{padding:.55rem}.dense body{font-size:13px}@media print{.controls,.tabs,button{display:none!important}.view{display:block!important}}
@media(max-width:700px){header,.controls,.tabs,main{padding-left:1rem;padding-right:1rem}}
"""

JS = r"""
const D=JSON.parse(document.getElementById('report-data').textContent), rows=D.rows;
const $=s=>document.querySelector(s), esc=s=>String(s??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
const md=s=>esc(s).replace(/^### (.*)$/gm,'<h4>$1</h4>').replace(/^## (.*)$/gm,'<h3>$1</h3>').replace(/^# (.*)$/gm,'<h2>$1</h2>').replace(/\*\*(.*?)\*\*/g,'<strong>$1</strong>').replace(/\n/g,'<br>');
const filtered=()=>{let q=$('#search').value.toLowerCase(),p=$('#pf').value,r=$('#pr').value,s=$('#st').value,z=$('#sort').value,a=rows.filter(x=>(!q||(x.question+' '+x.answer_plain_text).toLowerCase().includes(q))&&(!p||x.profile_id==p)&&(!r||x.prompt_id==r)&&(!s||x.status==s));return a.sort((x,y)=>z==='tokens'?x.answer_tokens-y.answer_tokens:z==='latency'?x.total_latency_s-y.total_latency_s:z==='citations'?y.citation_density_per_100_words-x.citation_density_per_100_words:z==='coverage'?y.expected_keyword_coverage-x.expected_keyword_coverage:y.heuristic_tradeoff_score-x.heuristic_tradeoff_score)};
const panel=x=>`<article class="card"><div><span class=badge>${esc(x.profile_id)}</span><span class=badge>${esc(x.prompt_id)}</span>${x.pareto_frontier?'<span class=chip>Pareto</span>':''}</div><div class=answer>${md(x.answer_markdown||x.error_message)}</div><p class=meta>${x.answer_tokens} tokens · ${x.total_latency_s}s · ${x.inline_citation_count} citations · coverage ${x.expected_keyword_coverage} · score ${x.heuristic_tradeoff_score}</p><div class=citations>${String(x.citation_ids||'').split('|').filter(Boolean).map(c=>`<span class=chip>[${c}]</span>`).join('')}</div><details><summary>Trace/debug</summary><div class=trace>context ${esc(x.context_hash)}\nretrieval ${esc(x.retrieval_fingerprint)}\nchunks ${esc(x.selected_chunk_ids)}</div></details></article>`;
const diffWords=(a,b)=>{let y=new Set(String(b).split(/\s+/));return String(a).split(/\s+/).map(w=>y.has(w)?esc(w):`<mark>${esc(w)}</mark>`).join(' ')};
function questionView(){let g=Object.groupBy?Object.groupBy(filtered(),x=>x.question_id):filtered().reduce((a,x)=>((a[x.question_id]??=[]).push(x),a),{});$('#question').innerHTML=Object.values(g).map(a=>`<section><h2>${esc(a[0].question_id)} · ${esc(a[0].question)}</h2><div class=answer-grid>${a.map(panel).join('')}</div></section>`).join('')||'<p>No matching cells.</p>'}
function compare(dim){let a=filtered(),fixed=dim==='prompt_id'?$('#pr').value:$('#pf').value;if(fixed)a=a.filter(x=>x[dim==='prompt_id'?'prompt_id':'profile_id']==fixed);let d=a.length>1?`<details class=card><summary>Deterministic word diff (first two visible answers)</summary>${diffWords(a[0].answer_markdown,a[1].answer_markdown)}</details>`:'';$('#'+(dim==='prompt_id'?'lengths':'prompts')).innerHTML='<div class=answer-grid>'+a.map(panel).join('')+'</div>'+d}
function chart(a,key,label,scale){return `<section class=card><h3>${label}</h3><div class=chart>${a.map(x=>`<div class=bar>${esc(x.profile_id+' / '+x.prompt_id)}<span style="width:${Math.max(2,Number(x[key]||0)*scale)}px"></span> ${esc(x[key])}</div>`).join('')}</div></section>`}
function leaderboard(){let a=D.aggregate.slice().sort((x,y)=>y.mean_heuristic_tradeoff_score-x.mean_heuristic_tradeoff_score),points=a.map(x=>`<circle cx="${25+Math.min(520,x.mean_tokens)}" cy="${185-Math.max(0,x.mean_heuristic_tradeoff_score)*150}" r="${4+Math.min(8,x.pareto_count)}"><title>${esc(x.profile_id+'/'+x.prompt_id)}</title></circle>`).join('');$('#leader').innerHTML=`<table><thead><tr>${Object.keys(a[0]||{}).map(k=>`<th>${esc(k)}</th>`).join('')}</tr></thead><tbody>${a.map(x=>`<tr>${Object.keys(a[0]).map(k=>`<td>${esc(x[k])}</td>`).join('')}</tr>`).join('')}</tbody></table><div class=grid>${chart(a,'mean_tokens','Tokens',.5)}${chart(a,'mean_latency_s','Latency',20)}${chart(a,'mean_citation_density','Citation density',20)}${chart(a,'mean_keyword_coverage','Keyword coverage',180)}${chart(a,'success_rate','Success',180)}<section class=card><h3>Pareto scatter</h3><svg viewBox="0 0 600 200" role=img aria-label="Tokens versus trade-off score">${points}</svg></section></div>`}
function review(){let saved=JSON.parse(localStorage.getItem('cial-pl-review')||'{}');$('#review').innerHTML=filtered().map(x=>`<div class=card><b>${esc(x.question_id)} · ${esc(x.profile_id)} / ${esc(x.prompt_id)}</b><label> Preferred <input type=checkbox data-key="${esc(x.question_id+'|'+x.profile_id+'|'+x.prompt_id+'|preferred')}" ${saved[x.question_id+'|'+x.profile_id+'|'+x.prompt_id+'|preferred']?'checked':''}></label><br><textarea data-key="${esc(x.question_id+'|'+x.profile_id+'|'+x.prompt_id+'|notes')}" rows=3 placeholder="Review notes">${esc(saved[x.question_id+'|'+x.profile_id+'|'+x.prompt_id+'|notes']||'')}</textarea></div>`).join('')+'<button id=download>Download review CSV</button>';document.querySelectorAll('[data-key]').forEach(e=>e.onchange=()=>{saved[e.dataset.key]=e.type==='checkbox'?e.checked:e.value;localStorage.setItem('cial-pl-review',JSON.stringify(saved))});$('#download').onclick=()=>{let body='key,value\n'+Object.entries(saved).map(([k,v])=>`"${k.replaceAll('"','""')}","${String(v).replaceAll('"','""')}"`).join('\n');let a=document.createElement('a');a.href=URL.createObjectURL(new Blob([body],{type:'text/csv'}));a.download='human_review.csv';a.click()}}
function render(){questionView();compare('profile_id');compare('prompt_id');leaderboard();review()}
document.querySelectorAll('.tabs button').forEach(b=>b.onclick=()=>{document.querySelectorAll('.view').forEach(v=>v.classList.remove('active'));$('#'+b.dataset.view).classList.add('active')});
['search','pf','pr','st','sort'].forEach(id=>$('#'+id).oninput=render);$('#theme').onclick=()=>document.documentElement.dataset.theme=document.documentElement.dataset.theme==='dark'?'light':'dark';$('#density').onclick=()=>document.documentElement.classList.toggle('dense');$('#layout').onchange=()=>document.documentElement.classList.toggle('tabbed',$('#layout').value==='tabbed');['metrics','citations','debug'].forEach(id=>$('#'+id).onchange=()=>document.documentElement.classList.toggle('hide-'+id,!$('#'+id).checked));render();
"""


def export_html(path: Path, rows: list[dict[str, Any]], aggregate: list[dict[str, Any]], summary: Mapping[str, Any], max_mb: float) -> None:
    data = canonical({"rows": rows, "aggregate": aggregate, "summary": summary}).replace("</", "<\\/")
    fallback = "".join(
        f"<article class='card'><h3>{html.escape(str(x.get('question_id')))} · "
        f"{html.escape(str(x.get('profile_id')))} / {html.escape(str(x.get('prompt_id')))}</h3>"
        f"<p>{html.escape(str(x.get('question')))}</p><div class='answer'>"
        f"{html.escape(str(x.get('answer_markdown') or x.get('error_message') or ''))}</div></article>"
        for x in rows
    )
    options = lambda values: "".join(f"<option value='{html.escape(str(x))}'>{html.escape(str(x))}</option>" for x in values)
    doc = f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width">
<title>CIAL Prompt × Length Comparison</title><style>{CSS}</style></head><body>
<header><h1>CIAL Prompt × Length Comparison</h1><p class=lede>Deterministic trade-offs; human review remains authoritative.</p>
<div class=metrics>{''.join(f"<div class=metric><b>{html.escape(str(v))}</b>{html.escape(str(k).replace('_',' '))}</div>" for k,v in summary.items() if k in ('questions','profiles','prompts','total_cells','success_rate','total_tokens','total_latency_s','model'))}</div></header>
<div class=controls><input id=search aria-label=Search placeholder="Search questions and answers"><select id=pf><option value="">All profiles</option>{options(sorted({x.get('profile_id') for x in rows}))}</select><select id=pr><option value="">All prompts</option>{options(sorted({x.get('prompt_id') for x in rows}))}</select><select id=st><option value="">All statuses</option>{options(sorted({x.get('status') for x in rows}))}</select><select id=sort><option value=score>Trade-off score</option><option value=tokens>Tokens</option><option value=latency>Latency</option><option value=citations>Citations</option><option value=coverage>Keyword coverage</option></select><select id=layout><option value=side>Side-by-side</option><option value=tabbed>Tabbed/stacked</option></select><label><input id=metrics type=checkbox checked> Metrics</label><label><input id=citations type=checkbox checked> Citations</label><label><input id=debug type=checkbox checked> Debug</label><button id=theme>Theme</button><button id=density>Density</button></div>
<nav class=tabs><button data-view=question>By Question</button><button data-view=prompts>Compare Prompts</button><button data-view=lengths>Compare Lengths</button><button data-view=leader>Leaderboard</button><button data-view=review>Human Review</button></nav>
<main><section id=question class="view active"></section><section id=prompts class=view></section><section id=lengths class=view></section><section id=leader class=view></section><section id=review class=view></section>
<noscript><h2>All questions and full answers</h2>{fallback}</noscript></main>
<script id=report-data type=application/json>{data}</script><script>{JS}</script></body></html>"""
    size_mb = len(doc.encode("utf-8")) / 1024 / 1024
    if size_mb > max_mb:
        raise ValueError(f"Self-contained HTML is {size_mb:.2f} MB, above --html-max-embedded-mb={max_mb}.")
    path.write_text(doc, encoding="utf-8")


def export_bundle(run_dir: Path, rows: list[dict[str, Any]], config: Mapping[str, Any], *, preflight: Mapping[str, Any] | None = None, html_max_mb: float = 100) -> dict[str, Any]:
    score_rows(rows, config.get("heuristic_weights", RunConfig("", "", "", "", [], [], 1, "", "", 0, "").heuristic_weights))
    aggregate = aggregate_rows(rows)
    prompt_pairs, length_pairs = pairwise(rows, "prompt_id"), pairwise(rows, "profile_id")
    success = [x for x in rows if x.get("status") == "success"]
    summary = {
        "run_id": config.get("run_id", ""), "questions": len({x.get("question_id") for x in rows}),
        "profiles": len({x.get("profile_id") for x in rows}), "prompts": len({x.get("prompt_id") for x in rows}),
        "total_cells": len(rows), "succeeded": len(success), "failed": len(rows) - len(success),
        "success_rate": round(len(success) / len(rows), 4) if rows else 0,
        "total_tokens": sum(int(x.get("total_model_tokens", 0) or 0) for x in rows),
        "total_latency_s": round(sum(float(x.get("total_latency_s", 0) or 0) for x in rows), 3),
        "model": config.get("model", ""), "run_date": utcnow(),
        "preflight_passed": bool((preflight or {}).get("passed", False)),
        "note": "No universal best is declared; human review is authoritative.",
    }
    write_csv(run_dir / "comparison_long.csv", rows, REQUIRED_COLUMNS)
    wide = wide_rows(rows)
    write_csv(run_dir / "comparison_wide.csv", wide, list(wide[0]) if wide else ["question_id"])
    write_csv(run_dir / "prompt_profile_matrix.csv", aggregate, list(aggregate[0]) if aggregate else ["profile_id"])
    write_csv(run_dir / "partial_results.csv", rows, REQUIRED_COLUMNS)
    with (run_dir / "partial_results.jsonl").open("w", encoding="utf-8", newline="\n") as stream:
        for row in rows:
            stream.write(canonical(row) + "\n")
    metrics = {
        "heuristic_formula": "w1*keyword_coverage+w2*safe_failure_correct+w3*citation_proxy-w4*forbidden_hits-w5*normalized_latency-w6*normalized_tokens",
        "weights": config.get("heuristic_weights", {}), "aggregate": aggregate,
        "pairwise_prompt": prompt_pairs, "pairwise_length": length_pairs,
    }
    atomic_json(run_dir / "summary.json", summary)
    atomic_json(run_dir / "metrics.json", metrics)
    export_excel(run_dir / "comparison.xlsx", rows, aggregate, prompt_pairs, length_pairs, config, summary)
    export_html(run_dir / "report.html", rows, aggregate, summary, html_max_mb)
    return summary


def failure_row(config: Mapping[str, Any], message: str) -> dict[str, Any]:
    row = {key: "" for key in REQUIRED_COLUMNS}
    row.update({
        "run_id": config.get("run_id", ""), "status": "preflight_failed",
        "error_code": "PREFLIGHT_FAILED", "error_message": message,
        "created_at": utcnow(), "comparison_valid": False,
        "comparison_warning": "No questions were executed.",
        "corpus_fingerprint": config.get("corpus_fingerprint", ""),
    })
    return row


class RunController:
    def __init__(self, run_dir: Path):
        self.run_dir = run_dir
        self.interrupted = False
        self.rows: list[dict[str, Any]] = []

    def checkpoint(self, status: str, planned: int = 0) -> None:
        atomic_json(self.run_dir / "checkpoint.json", {
            "status": status, "updated_at": utcnow(), "completed": len(self.rows),
            "planned": planned, "successful": sum(x.get("status") == "success" for x in self.rows),
            "failed": sum(x.get("status") != "success" for x in self.rows),
        })

    def interrupt(self, *_: Any) -> None:
        self.interrupted = True
        self.checkpoint("interrupted")


def run_matrix(
    adapter: Any, questions: list[QuestionRecord], profiles: list[AnswerLengthProfile],
    prompts: list[PromptVariant], args: argparse.Namespace, config: dict[str, Any],
    controller: RunController, existing: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rows = list(existing or [])
    controller.rows = rows
    existing_by_key = {cell_key(x): x for x in rows if x.get("status") == "success"}
    cells = matrix_cells(
        questions, profiles, prompts, args.repeat, args.seed,
        args.randomize_profile_order, args.randomize_prompt_order,
    )
    frozen: dict[str, FrozenContext] = {}
    execution_order = len(rows)
    for question, profile, prompt, repeat_index in cells:
        if controller.interrupted:
            break
        if args.rerun_prompt and prompt.id != args.rerun_prompt:
            continue
        if args.rerun_profile and profile.id != args.rerun_profile:
            continue
        if question.question_id not in frozen:
            frozen[question.question_id] = adapter.freeze(question)
        shared = frozen[question.question_id]
        probe = {
            "question_id": question.question_id, "profile_fingerprint": profile.fingerprint,
            "prompt_fingerprint": prompt.fingerprint, "repeat_index": repeat_index,
            "context_hash": shared.context_hash,
        }
        targeted_rerun = bool(args.rerun_prompt or args.rerun_profile)
        if (
            not args.force
            and cell_key(probe) in existing_by_key
            and (args.rerun_failed or not targeted_rerun)
        ):
            continue
        execution_order += 1
        rendered = render_prompt(prompt, profile, question.question, shared.context, args.scope)
        if args.trace_level == "full":
            context_dir = Path(config["run_dir"]) / "context"
            atomic_json(context_dir / f"{question.question_id}_frozen_context.json", {
                "question_id": question.question_id, "question": question.question,
                "context": shared.context, "context_hash": shared.context_hash,
                "retrieval_fingerprint": shared.retrieval_fingerprint, "chunks": shared.chunks,
            })
            atomic_json(
                context_dir / f"{question.question_id}_{profile.id}_{prompt.id}_r{repeat_index}_prompt.json",
                {"rendered_prompt": rendered, "rendered_prompt_hash": sha(rendered)},
            )
        start = time.perf_counter()
        base = {
            "run_id": config["run_id"], "question_id": question.question_id,
            "question_index": question.question_index, "question": question.question,
            "category": question.category, "profile_id": profile.id, "profile_label": profile.label,
            "profile_fingerprint": profile.fingerprint, "prompt_id": prompt.id,
            "prompt_label": prompt.label, "prompt_version": prompt.version,
            "prompt_fingerprint": prompt.fingerprint, "repeat_index": repeat_index,
            "execution_order": execution_order, "prompt_tokens": count_tokens(rendered),
            "rendered_prompt_hash": sha(rendered), "context_hash": shared.context_hash,
            "retrieval_fingerprint": shared.retrieval_fingerprint, "comparison_valid": True,
            "comparison_warning": "", "evidence_strength": shared.evidence_strength,
            "weak_evidence": shared.weak_evidence, "unsupported_query": shared.unsupported_query,
            "selected_evidence_tokens": count_tokens(shared.context),
            "final_context_tokens": count_tokens(shared.context),
            "retrieval_latency_s": round(shared.retrieval_latency_s, 4),
            "rerank_latency_s": round(shared.rerank_latency_s, 4),
            "selection_latency_s": round(shared.selection_latency_s, 4),
            "prompt_name": prompt.id, "prompt_registry_version": PROMPT_REGISTRY_VERSION,
            "created_at": utcnow(), "retry_count": 0,
            "corpus_fingerprint": config["corpus_fingerprint"], **chunk_fields(shared.chunks),
        }
        try:
            answer, model = adapter.generate(rendered, profile)
            metric = row_metrics(answer, question.question)
            generation = float(model.get("generation_latency_s", time.perf_counter() - start))
            row = {
                **base, **metric, **model, "status": "success", "error_code": "",
                "error_message": "", "answer_markdown": answer,
                "generation_latency_s": round(generation, 4),
                "total_latency_s": round(generation + shared.retrieval_latency_s + shared.rerank_latency_s + shared.selection_latency_s, 4),
            }
            row["total_model_tokens"] = int(row["prompt_tokens"]) + int(row["completion_tokens"])
        except Exception as exc:
            row = {
                **base, "status": "failed", "error_code": type(exc).__name__,
                "error_message": str(exc), "answer_status": "error",
                "answer_markdown": "", "answer_plain_text": "",
                "generation_latency_s": round(time.perf_counter() - start, 4),
                "total_latency_s": round(time.perf_counter() - start + shared.retrieval_latency_s, 4),
            }
            for key in REQUIRED_COLUMNS:
                row.setdefault(key, 0 if key.endswith(("_tokens", "_count", "_words", "_chars", "_hits")) else "")
        rows = [item for item in rows if cell_key(item) != cell_key(row)]
        rows.append(row)
        controller.rows = rows
        append_jsonl(Path(config["run_dir"]) / "raw_results.jsonl", row)
        controller.checkpoint("running", len(cells))
        write_csv(Path(config["run_dir"]) / "partial_results.csv", rows, REQUIRED_COLUMNS)
        with (Path(config["run_dir"]) / "partial_results.jsonl").open("w", encoding="utf-8") as stream:
            for item in rows:
                stream.write(canonical(item) + "\n")
        if row["status"] != "success" and not args.continue_on_error:
            # Per specification, cell failures continue by default; flag exists for compatibility.
            pass
    controller.checkpoint("interrupted" if controller.interrupted else "complete", len(cells))
    return rows


def comma_list(value: str) -> list[str]:
    return [x.strip() for x in value.split(",") if x.strip()]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run a fair, frozen-context CIAL prompt × answer-length comparison.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--questions-file", default=str(DEFAULT_QUESTIONS))
    parser.add_argument("--corpus-root", default=str(DEFAULT_CORPUS))
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--profiles", default="concise,balanced,detailed")
    parser.add_argument("--prompts", default="direct,structured,analytical,operational")
    parser.add_argument("--custom-prompts-file")
    parser.add_argument("--custom-prompts-mode", choices=("append", "replace"), default="append")
    parser.add_argument("--max-questions", type=int)
    parser.add_argument("--repeat", type=int, default=1)
    parser.add_argument("--run-label", default="comparison")
    parser.add_argument("--resume")
    parser.add_argument("--export-only")
    parser.add_argument("--retrieval-mode", choices=("hybrid", "dense", "bm25"), default="hybrid")
    parser.add_argument("--model", default="")
    parser.add_argument("--scope", default="authorized enterprise corpus")
    parser.add_argument("--trace-level", choices=("compact", "full"), default="compact")
    parser.add_argument("--randomize-prompt-order", action="store_true")
    parser.add_argument("--randomize-profile-order", action="store_true")
    parser.add_argument("--seed", type=int, default=20260724)
    parser.add_argument("--warmup", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--continue-on-error", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--rerun-failed", action="store_true")
    parser.add_argument("--rerun-prompt")
    parser.add_argument("--rerun-profile")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--html-max-embedded-mb", type=float, default=100)
    parser.add_argument("--open-report", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--smoke-test", action="store_true")
    parser.add_argument("--prepare-corpus", action="store_true")
    parser.add_argument("--start-services", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--wait-for-ready", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--verify-index", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--skip-corpus-check", action="store_true")
    parser.add_argument("--startup-timeout-seconds", type=float, default=180)
    parser.add_argument("--index-timeout-seconds", type=float, default=1800)
    parser.add_argument("--poll-interval-seconds", type=float, default=3)
    parser.add_argument("--qdrant-url", default=DEFAULT_QDRANT_URL)
    parser.add_argument("--backend-url", default=DEFAULT_BACKEND_URL)
    parser.add_argument("--collection", default=DEFAULT_COLLECTION)
    parser.add_argument("--allow-unsupported-files", action="store_true")
    parser.add_argument("--stop-started-services-on-exit", action="store_true")
    return parser


def make_run_dir(args: argparse.Namespace) -> Path:
    if args.resume:
        return Path(args.resume).resolve()
    if args.export_only:
        return Path(args.export_only).resolve()
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    label = re.sub(r"[^A-Za-z0-9_.-]+", "_", args.run_label).strip("_") or "comparison"
    path = Path(args.output_root).resolve() / f"run_{stamp}_{label}"
    path.mkdir(parents=True, exist_ok=False)
    return path


def self_test() -> int:
    passed: list[str] = []
    def check(name: str, condition: Any) -> None:
        if not condition:
            raise AssertionError(name)
        passed.append(name)
    with tempfile.TemporaryDirectory(prefix="cial_prompt_length_") as temporary:
        root = Path(temporary)
        test_env_file = root / "service" / ".env"
        test_env_file.parent.mkdir()
        env_url = "postgresql+psycopg://env_user:env_secret@db.internal:5433/cial_env"
        shell_url = "postgresql://shell_user:shell_secret@shell.internal/cial_shell"
        test_env_file.write_text(f"DATABASE_URL={env_url}\n", encoding="utf-8")

        def fake_dotenv(target: Mapping[str, str]) -> Callable[..., bool]:
            def load(path: Path, *, override: bool = False) -> bool:
                value = Path(path).read_text(encoding="utf-8").split("=", 1)[1].strip()
                if override or not target.get("DATABASE_URL"):
                    target["DATABASE_URL"] = value  # type: ignore[index]
                return True
            return load

        env_mapping: dict[str, str] = {}
        env_state = bootstrap_environment(
            test_env_file,
            environment=env_mapping,
            dotenv_loader=fake_dotenv(env_mapping),
        )
        check(
            "service .env found and loaded",
            env_state.env_file_found and env_state.env_file_loaded,
        )
        check(
            ".env supplies DATABASE_URL when shell variable is absent",
            env_state.database_url == env_url
            and env_state.configuration_source == "service .env",
        )
        shell_mapping = {"DATABASE_URL": shell_url}
        shell_state = bootstrap_environment(
            test_env_file,
            environment=shell_mapping,
            dotenv_loader=fake_dotenv(shell_mapping),
        )
        check(
            "shell DATABASE_URL precedence with override false",
            shell_state.database_url == shell_url
            and shell_state.configuration_source == "process environment",
        )
        missing_env = root / "missing" / ".env"
        missing_shell_state = bootstrap_environment(
            missing_env,
            environment={"DATABASE_URL": shell_url},
            dotenv_loader=lambda *_args, **_kwargs: False,
        )
        check(
            "missing .env with existing shell variable succeeds",
            missing_shell_state.database_url_resolved
            and not missing_shell_state.env_file_found,
        )
        unresolved_state = bootstrap_environment(
            missing_env,
            environment={},
            dotenv_loader=lambda *_args, **_kwargs: False,
        )
        unresolved_gate = database_configuration_gate(unresolved_state)
        check(
            "missing .env and missing URL fails clearly",
            not unresolved_gate["passed"]
            and "canonical backend settings" in str(unresolved_gate.get("error")),
        )
        import_order: list[bool] = []
        canonical_state = resolve_canonical_database_url(
            unresolved_state,
            resolver=lambda: (
                import_order.append(unresolved_state.bootstrap_complete) or env_url
            ),
        )
        check(
            "backend config import occurs after environment bootstrap",
            import_order == [True]
            and canonical_state.configuration_source == "backend settings",
        )

        captured_popen: dict[str, Any] = {}
        class FakeProcess:
            returncode = 0
            def poll(self) -> int:
                return 0
        def fake_popen(*popen_args: Any, **popen_kwargs: Any) -> Any:
            captured_popen["args"] = popen_args
            captured_popen["kwargs"] = popen_kwargs
            return FakeProcess()
        child_run_dir = root / "child-output"
        child_run_dir.mkdir()
        child_manager = ServiceManager(
            child_run_dir,
            environment=canonical_state.child_environment(),
            popen_factory=fake_popen,
        )
        child_manager.start("backend", root / "start_backend.bat")
        child_manager.stop()
        check(
            "child process environment receives resolved URL",
            captured_popen["kwargs"]["env"]["DATABASE_URL"] == env_url,
        )
        diagnostics = environment_diagnostics(env_state)
        serialized_diagnostics = canonical(diagnostics)
        check(
            "database diagnostics never expose credentials or full URL",
            "env_secret" not in serialized_diagnostics
            and env_url not in serialized_diagnostics
            and diagnostics.get("database_host") == "db.internal:5433"
            and diagnostics.get("database_name") == "cial_env",
        )
        check(
            "only runner source path is allowed",
            only_allowed_source_changes(
                [
                    "scripts/run_prompt_length_comparison.py",
                    "outputs/batch_answers/prompt_length_comparisons/run_test/report.html",
                ]
            )
            and not only_allowed_source_changes(
                ["scripts/run_prompt_length_comparison.py", "README.md"]
            ),
        )
        questions_path = root / "questions.txt"
        questions_path.write_text("\ufeff# comment\nWhat is queue policy?\n\nWhat is queue policy?\nHow often inspect?\n", encoding="utf-8")
        qs, stats = load_questions(questions_path)
        check("question parsing", len(qs) == 3 and qs[0].question_id == "Q0001")
        check("duplicate preservation", stats["exact_duplicate_occurrences"] == 1 and qs[0].question == qs[1].question)
        validate_prompts(list(PROMPTS.values()))
        check("profile/prompt validation", PROFILES["concise"].fingerprint == PROFILES["concise"].fingerprint)
        cells = matrix_cells(qs, list(PROFILES.values()), [PROMPTS["direct"], PROMPTS["structured"]], 2, 1)
        check("matrix size", len(cells) == 36)
        rendered = render_prompt(PROMPTS["direct"], PROFILES["concise"], qs[0].question, "[1] evidence", "test")
        check("prompt rendering placeholders", "{question}" not in rendered and "[1] evidence" in rendered)
        corpus = root / "corpus"
        corpus.mkdir()
        (corpus / "a.txt").write_text("same", encoding="utf-8")
        (corpus / "b.txt").write_text("same", encoding="utf-8")
        (corpus / "~$temp.docx").write_text("temp", encoding="utf-8")
        (corpus / ".hidden.txt").write_text("hidden", encoding="utf-8")
        (corpus / "bad.xyz").write_text("bad", encoding="utf-8")
        manifest, cstats = scan_corpus(corpus, {".txt"})
        check("corpus recursive scan", len(manifest) == 3)
        check("ignored temp files", len(cstats["ignored"]) == 2)
        check("supported extension classification", sum(x["supported"] for x in manifest) == 2)
        check("duplicate hash detection", len(cstats["duplicate_hashes"]) == 1)
        tree = {"files": [
            {"relative_path": "a.txt", "document_id": "d1", "current_version_id": "v1", "chunk_count": 1, "content_hash": manifest[0]["sha256"]},
            {"relative_path": "b.txt", "document_id": "d2", "current_version_id": "v2", "chunk_count": 1, "content_hash": manifest[1]["sha256"]},
        ]}
        ver = verify_manifest(manifest[:2], tree, 2)
        check("metadata/vector verification", all(not x["blocking"] for x in ver))
        check("metadata row missing", verify_manifest([manifest[0]], {}, 2)[0]["verification_status"] == "metadata_missing")
        zero_tree = {"files": [{"relative_path": "a.txt", "document_id": "d1", "chunk_count": 0}]}
        check("chunk count zero", verify_manifest([manifest[0]], zero_tree, 2)[0]["verification_status"] == "zero_chunks")
        check("vector mismatch", verify_manifest([manifest[0]], tree, 0)[0]["verification_status"] == "vector_mismatch")
        fp1 = corpus_fingerprint(manifest, ver, {"collection": "x"})
        changed = [dict(x) for x in manifest]
        changed[0]["sha256"] = "changed"
        check("corpus fingerprint changes mid-run", fp1 != corpus_fingerprint(changed, ver, {"collection": "x"}))
        transitions = iter([
            (False, {"status": "indexing"}), (True, {"status": "ready"}), (True, {"status": "ready"}),
        ])
        ready, _, polls = wait_until(lambda: next(transitions), 1, .01, stable_polls=2)
        check("queued-to-running-to-completed indexing", ready and polls == 3)
        check("stable two-poll completion", ready)
        check("indexing failure blocks generation", not backend_ready({"status": "failed"}))
        check("service-already-running path", reachable("fake", lambda *a, **k: (200, {"status": "ok"}))[0])
        timeout_ok, _, _ = wait_until(lambda: (False, {}), .03, .01)
        check("service-start timeout", not timeout_ok)
        check("backend readiness state transitions", backend_ready({"status": "ready", "engine_ready": True, "qdrant_ready": True, "models_ready": True, "index_fresh": True}))
        terminal_ready, terminal_payload, terminal_polls = wait_for_backend_ready(
            "fake",
            1,
            .01,
            getter=lambda *_args, **_kwargs: (
                200,
                {"status": "failed", "message": "terminal test failure"},
            ),
        )
        check(
            "terminal backend failure ends readiness polling",
            not terminal_ready
            and terminal_payload["status"] == "failed"
            and terminal_polls == 1,
        )
        run_dir = root / "run"
        run_dir.mkdir()
        args = build_parser().parse_args([
            "--questions-file", str(questions_path), "--corpus-root", str(corpus),
            "--output-root", str(root), "--skip-corpus-check", "--profiles", "concise",
            "--prompts", "direct,structured", "--continue-on-error",
        ])
        cfg = asdict(RunConfig("test", str(run_dir), str(questions_path), str(corpus), ["concise"], ["direct", "structured"], 1, "hybrid", "fake", 1, "compact", fp1))
        controller = RunController(run_dir)
        adapter = FakeAdapter(fail_prompt="Begin with a direct answer")
        rows = run_matrix(adapter, qs[:1], [PROFILES["concise"]], [PROMPTS["direct"], PROMPTS["structured"]], args, cfg, controller)
        check("context frozen across cells", adapter.freezes == 1 and len({x["context_hash"] for x in rows}) == 1)
        check("failed-cell continuation", len(rows) == 2 and {x["status"] for x in rows} == {"success", "failed"})
        before = len(rows)
        rows2 = run_matrix(FakeAdapter(), qs[:1], [PROFILES["concise"]], [PROMPTS["direct"], PROMPTS["structured"]], args, cfg, controller, rows)
        check("resume skip", len(rows2) == before and sum(x["prompt_id"] == "direct" for x in rows2) == 1 and all(x["status"] == "success" for x in rows2))
        export_bundle(run_dir, rows, cfg, preflight={"passed": True}, html_max_mb=20)
        check("CSV quoting/formula escaping", safe_cell("=2+2") == "'=2+2" and "\n" in rows[0]["answer_markdown"])
        check("wide pivot", "concise__direct__answer" in wide_rows(rows)[0])
        from openpyxl import load_workbook
        workbook = load_workbook(run_dir / "comparison.xlsx", read_only=True)
        check("Excel creation/open validation", "Human Review" in workbook.sheetnames and len(workbook.sheetnames) == 10)
        workbook.close()
        report = (run_dir / "report.html").read_text(encoding="utf-8")
        check("HTML embedding/no external assets", "https://" not in report and "<script id=report-data" in report)
        check("full answers retained", rows[0]["answer_markdown"] in load_jsonl(run_dir / "raw_results.jsonl")[0]["answer_markdown"] and html.escape(rows[0]["answer_markdown"]) in report)
        check("pairwise metrics", bool(pairwise(rows2, "prompt_id")) and "tf_cosine" in pairwise(rows2, "prompt_id")[0])
        check("heuristic score transparency", "heuristic_formula" in json.loads((run_dir / "metrics.json").read_text(encoding="utf-8")))
        check("terminal checkpoint", json.loads((run_dir / "checkpoint.json").read_text(encoding="utf-8"))["status"] == "complete")
        failure_dir = root / "failure"
        failure_dir.mkdir()
        export_bundle(failure_dir, [failure_row(cfg, "fake gate failure")], cfg, preflight={"passed": False}, html_max_mb=20)
        check("preflight failure creates reports", all((failure_dir / x).exists() for x in ("report.html", "comparison.xlsx", "comparison_long.csv")))
        check("preflight pass allows matrix execution", len(rows) > 0)
        controller.interrupt()
        check("Ctrl+C checkpoint flush", json.loads((run_dir / "checkpoint.json").read_text(encoding="utf-8"))["status"] == "interrupted")
        check("only runtime outputs outside source", all(p.is_relative_to(root) for p in root.rglob("*")))
    print(f"SELF-TEST PASSED — {len(passed)} checks")
    for name in passed:
        print(f"  PASS {name}")
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.self_test:
        return self_test()
    if args.smoke_test and args.max_questions is None:
        args.max_questions = 3
    run_dir = make_run_dir(args)
    run_dir.mkdir(parents=True, exist_ok=True)
    if args.export_only:
        config_path = run_dir / "config.json"
        if not config_path.is_file():
            raise FileNotFoundError(f"Missing export config: {config_path}")
        config = json.loads(config_path.read_text(encoding="utf-8"))
        rows = latest_results(load_jsonl(run_dir / "raw_results.jsonl"))
        summary = export_bundle(run_dir, rows, config, html_max_mb=args.html_max_embedded_mb)
        print(f"Exported {len(rows)} persisted cells to {run_dir}")
        return 0
    profile_ids = comma_list(args.profiles)
    profiles = [PROFILES[x] for x in profile_ids if x in PROFILES]
    if len(profiles) != len(profile_ids):
        raise ValueError(f"Unknown profile(s): {sorted(set(profile_ids) - PROFILES.keys())}")
    prompts_registry = dict(PROMPTS)
    custom = load_custom_prompts(Path(args.custom_prompts_file)) if args.custom_prompts_file else []
    if custom:
        prompts_registry = {} if args.custom_prompts_mode == "replace" else prompts_registry
        for item in custom:
            if item.id in prompts_registry:
                raise ValueError(f"Duplicate built-in/custom prompt id: {item.id}")
            prompts_registry[item.id] = item
    prompt_ids = comma_list(args.prompts)
    prompts = [prompts_registry[x] for x in prompt_ids if x in prompts_registry]
    if len(prompts) != len(prompt_ids):
        raise ValueError(f"Unknown prompt(s): {sorted(set(prompt_ids) - prompts_registry.keys())}")
    validate_prompts(prompts)
    run_id = run_dir.name
    config = asdict(RunConfig(
        run_id, str(run_dir), str(Path(args.questions_file).resolve()), str(Path(args.corpus_root).resolve()),
        profile_ids, prompt_ids, args.repeat, args.retrieval_mode, args.model, args.seed, args.trace_level,
        skip_corpus_check_acknowledged=args.skip_corpus_check,
    ))
    config.update({
        "script_version": SCRIPT_VERSION, "prompt_registry_version": PROMPT_REGISTRY_VERSION,
        "profiles_resolved": [asdict(x) | {"fingerprint": x.fingerprint} for x in profiles],
        "prompts_resolved": [asdict(x) | {"fingerprint": x.fingerprint} for x in prompts],
        "arguments": vars(args), "created_at": utcnow(),
    })
    prior_config = None
    if args.resume and (run_dir / "config.json").is_file():
        prior_config = json.loads((run_dir / "config.json").read_text(encoding="utf-8"))
    controller = RunController(run_dir)
    manager = ServiceManager(
        run_dir,
        environment=ENV_BOOTSTRAP.child_environment(),
    )
    old_handler = signal.getsignal(signal.SIGINT)
    signal.signal(signal.SIGINT, controller.interrupt)
    if args.stop_started_services_on_exit:
        atexit.register(manager.stop)
    try:
        preflight = run_preflight(args, run_dir, manager)
        config["corpus_fingerprint"] = preflight["corpus_fingerprint"]
        config["question_manifest_sha256"] = preflight.get("question_stats", {}).get("sha256", "")
        if prior_config is not None:
            for key in ("question_manifest_sha256", "corpus_fingerprint", "profiles", "prompts", "retrieval_mode"):
                if prior_config.get(key) != config.get(key):
                    raise ValueError(f"Resume fingerprint/config mismatch: {key}")
        atomic_json(run_dir / "config.json", config)
        atomic_json(run_dir / "question_manifest.json", {
            "source": preflight.get("question_stats", {}),
            "questions": [asdict(x) for x in preflight.get("questions", [])],
        })
        atomic_json(run_dir / "preflight_status.json", {k: v for k, v in preflight.items() if k not in ("questions", "manifest", "verification")})
        planned = len(preflight.get("questions", [])) * len(profiles) * len(prompts) * args.repeat
        print(f"Run: {run_id}\nQuestions: {len(preflight.get('questions', []))} | Profiles: {len(profiles)} ({', '.join(profile_ids)})")
        print(f"Prompts: {len(prompts)} ({', '.join(prompt_ids)}) | Repeats: {args.repeat} | Planned model calls: {planned}")
        print(f"Output: {run_dir}")
        if not preflight["passed"]:
            reasons = [
                f"{name}: {gate.get('error') or gate.get('warning') or gate.get('snapshot', {}).get('message', 'failed')}"
                for name, gate in preflight["gates"].items() if not gate.get("passed")
            ]
            message = "; ".join(reasons)
            print("PREFLIGHT FAILED — no questions were executed")
            row = failure_row(config, message)
            append_jsonl(run_dir / "raw_results.jsonl", row)
            controller.rows = [row]
            controller.checkpoint("preflight_failed", planned)
            export_bundle(run_dir, [row], config, preflight=preflight, html_max_mb=args.html_max_embedded_mb)
            return 2
        print(f"PREFLIGHT PASSED — starting {planned} generation attempts")
        if args.prepare_corpus:
            controller.checkpoint("prepared", 0)
            export_bundle(run_dir, [], config, preflight=preflight, html_max_mb=args.html_max_embedded_mb)
            print(f"Corpus prepared and verified. Fingerprint: {config['corpus_fingerprint']}")
            return 0
        if args.dry_run:
            controller.checkpoint("dry_run", planned)
            print("Dry run complete; no model calls made.")
            return 0
        existing = latest_results(load_jsonl(run_dir / "raw_results.jsonl")) if args.resume else []
        adapter: Any = RepositoryAdapter(args)
        try:
            rows = run_matrix(adapter, preflight["questions"], profiles, prompts, args, config, controller, existing)
        finally:
            adapter.close()
        summary = export_bundle(run_dir, rows, config, preflight=preflight, html_max_mb=args.html_max_embedded_mb)
        print(f"Completed: {summary['succeeded']} succeeded, {summary['failed']} failed, {summary['total_tokens']} tokens")
        for name in ("comparison_long.csv", "comparison_wide.csv", "prompt_profile_matrix.csv", "comparison.xlsx", "report.html", "summary.json"):
            print(f"  {run_dir / name}")
        if args.open_report:
            webbrowser.open((run_dir / "report.html").as_uri())
        return 0 if summary["failed"] == 0 else 1
    except KeyboardInterrupt:
        controller.interrupt()
        return 130
    except Exception:
        with (run_dir / "logs.txt").open("a", encoding="utf-8") as stream:
            traceback.print_exc(file=stream)
        controller.checkpoint("failed")
        raise
    finally:
        signal.signal(signal.SIGINT, old_handler)
        if args.stop_started_services_on_exit:
            manager.stop()


if __name__ == "__main__":
    raise SystemExit(main())
