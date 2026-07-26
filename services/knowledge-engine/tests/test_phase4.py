from __future__ import annotations

import csv
import json
import io
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from unittest.mock import patch

from openpyxl import load_workbook

from cial_knowledge_os.batch_qa import (
    CSV_COLUMNS,
    PHASE2_CSV_COLUMNS,
    PHASE3_CSV_COLUMNS,
    PHASE4_CSV_COLUMNS,
)
from cial_knowledge_os.config import (
    KnowledgeOSConfig,
    Phase2Config,
    Phase3Config,
    Phase4Config,
)
from cial_knowledge_os.context_builder import INSUFFICIENT_EVIDENCE_RESPONSE
from cial_knowledge_os.evidence_quality import EvidenceQualityScorer
from cial_knowledge_os.evidence_selector import EvidenceSelector
from cial_knowledge_os.llm import build_grounded_prompt
from cial_knowledge_os.phase4_pipeline import (
    Phase4RAGPipeline,
    UNSUPPORTED_QUERY_RESPONSE,
)
from cial_knowledge_os.phase4_runner import Phase4Runner
from cial_knowledge_os.phase4_trace import Phase4Trace, phase4_diagnostics
from cial_knowledge_os.reranker import (
    CrossEncoderReranker,
    MockReranker,
    resolve_reranker_device,
)
from cial_knowledge_os.token_budget import TokenBudgetManager


class _CharacterTokenizer:
    def encode(self, text: str, **_: Any) -> list[int]:
        return [ord(character) for character in text]

    def decode(self, values: list[int], **_: Any) -> str:
        return "".join(chr(value) for value in values)


class _StaticRetriever:
    def __init__(self, name: str, results: list[dict[str, Any]]) -> None:
        self.name = name
        self.results = results

    def retrieve(self, query: str, *, top_k: int) -> list[dict[str, Any]]:
        return [dict(item) for item in self.results[:top_k]]


class _CitingLLM:
    def invoke(self, prompt: str) -> str:
        self.prompt = prompt
        return "**Use the selected control** [1]."


class _RefusingLLM:
    def invoke(self, prompt: str) -> str:
        self.prompt = prompt
        return INSUFFICIENT_EVIDENCE_RESPONSE


class _FlakyLLM:
    def __init__(self, failures: int) -> None:
        self.failures = failures
        self.calls = 0

    def invoke(self, prompt: str) -> str:
        self.prompt = prompt
        self.calls += 1
        if self.calls <= self.failures:
            raise RuntimeError(
                "model runner has unexpectedly stopped: std::bad_alloc "
                "(status code: 500)"
            )
        return "**Recovered grounded answer** [1]."


class _FakeCrossEncoder:
    def __init__(self) -> None:
        self.calls: list[tuple[Any, ...]] = []
        self.device = "cuda:0"

    def parameters(self):
        yield SimpleNamespace(dtype="torch.float32")

    def predict(self, pairs, **kwargs):
        self.calls.append((pairs, kwargs))
        return [0.2, 0.9]


class _ReadyPipeline:
    def __init__(self, pipeline: Phase4RAGPipeline) -> None:
        self.pipeline = pipeline
        self.config = pipeline.config
        self.metrics = pipeline.metrics
        self.token_manager = pipeline.token_manager

    @property
    def is_ready_for_answering(self) -> bool:
        return True

    def answer(self, question: str):
        return self.pipeline.answer(question)


def _candidate(
    chunk: str,
    text: str,
    *,
    source: str,
    page: int,
    score: float = 0.5,
) -> dict[str, Any]:
    return {
        "id": chunk,
        "text": text,
        "score": score,
        "source": Path(source).name,
        "page_number": page,
        "chunk_id": chunk,
        "metadata": {
            "source": source,
            "file_name": Path(source).name,
            "page_number": page,
            "chunk_id": chunk,
            "chunk_index": page,
        },
    }


class RerankerAndSelectionTests(unittest.TestCase):
    def test_cross_encoder_interface_preserves_phase3_scores(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker(
            "local-test-model",
            model=model,
            batch_size=2,
        )
        candidates = [
            _candidate("a", "weak", source="C:/docs/a.pdf", page=1, score=0.7),
            _candidate("b", "strong", source="C:/docs/b.pdf", page=2, score=0.4),
        ]

        result = reranker.rerank("Which control?", candidates)

        self.assertEqual([item["chunk_id"] for item in result.candidates], ["b", "a"])
        self.assertEqual(result.candidates[0]["reranker_score"], 0.9)
        self.assertEqual(result.candidates[0]["score"], 0.4)
        self.assertEqual(result.candidates[0]["original_rrf_rank"], 2)
        self.assertEqual(model.calls[0][1]["batch_size"], 2)
        self.assertEqual(len(model.calls), 1)
        self.assertEqual(
            reranker.last_rerank_metrics["reranker_candidate_count"],
            2,
        )
        self.assertEqual(
            reranker.last_rerank_metrics["reranker_batch_size"],
            2,
        )
        self.assertEqual(
            reranker.last_rerank_metrics["reranker_device"],
            "cuda:0",
        )
        self.assertEqual(
            reranker.last_rerank_metrics["reranker_dtype"],
            "torch.float32",
        )

    def test_reranker_auto_uses_cuda_when_available(self) -> None:
        with patch("torch.cuda.is_available", return_value=True):
            self.assertEqual(resolve_reranker_device("auto"), "cuda")

    def test_reranker_auto_falls_back_to_cpu_without_cuda(self) -> None:
        with patch("torch.cuda.is_available", return_value=False):
            self.assertEqual(resolve_reranker_device("auto"), "cpu")

    def test_explicit_cpu_reranker_remains_on_cpu(self) -> None:
        self.assertEqual(resolve_reranker_device("cpu"), "cpu")

    def test_reranker_warmup_initializes_inference_only_once(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker(
            "local-test-model",
            model=model,
            batch_size=16,
        )

        reranker.warm()
        reranker.warm()

        self.assertEqual(len(model.calls), 1)
        self.assertEqual(model.calls[0][1]["batch_size"], 1)
        diagnostics = reranker.runtime_diagnostics()
        self.assertTrue(diagnostics["reranker_model_loaded"])
        self.assertTrue(diagnostics["reranker_warmed"])
        self.assertIsNotNone(diagnostics["reranker_warm_duration_ms"])

    def test_reranker_loads_from_cache_before_considering_download(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker("approved/reranker")
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            return_value=model,
        ) as cross_encoder, redirect_stdout(output):
            loaded = reranker._load_model()

        self.assertIs(loaded, model)
        self.assertEqual(reranker.load_source, "cache")
        self.assertTrue(cross_encoder.call_args.kwargs["local_files_only"])
        self.assertIn("loaded from local Hugging Face cache", output.getvalue())

    def test_reranker_downloads_once_after_cache_miss(self) -> None:
        model = _FakeCrossEncoder()
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=False,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=[OSError("cache miss"), model],
        ) as cross_encoder, redirect_stdout(output):
            loaded = reranker._load_model()

        self.assertIs(loaded, model)
        self.assertEqual(reranker.load_source, "download")
        self.assertEqual(
            [
                call.kwargs["local_files_only"]
                for call in cross_encoder.call_args_list
            ],
            [True, False],
        )
        self.assertIn(
            'Downloading reranker model "approved/reranker"',
            output.getvalue(),
        )
        self.assertIn(
            "Reranker downloaded and cached successfully",
            output.getvalue(),
        )

    def test_enterprise_offline_mode_never_attempts_download(self) -> None:
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=True,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=OSError("cache miss"),
        ) as cross_encoder, redirect_stdout(output):
            with self.assertRaisesRegex(
                RuntimeError,
                'Configured reranker model: "approved/reranker"',
            ) as context:
                reranker._load_model()

        self.assertEqual(cross_encoder.call_count, 1)
        self.assertTrue(cross_encoder.call_args.kwargs["local_files_only"])
        self.assertIn("Local-only mode: enabled", str(context.exception))
        self.assertIn("MockReranker", str(context.exception))
        self.assertIn(
            "Download skipped because enterprise offline mode is enabled",
            output.getvalue(),
        )

    def test_download_failure_message_is_actionable(self) -> None:
        reranker = CrossEncoderReranker(
            "approved/reranker",
            local_files_only=False,
        )
        output = io.StringIO()

        with patch(
            "sentence_transformers.CrossEncoder",
            side_effect=[
                OSError("cache miss"),
                OSError("network unavailable"),
            ],
        ), redirect_stdout(output):
            with self.assertRaises(RuntimeError) as context:
                reranker._load_model()

        message = str(context.exception)
        self.assertIn('Configured reranker model: "approved/reranker"', message)
        self.assertIn("Local-only mode: disabled", message)
        self.assertIn("stage it manually", message)
        self.assertIn("MockReranker", message)

    def test_mock_reranker_is_deterministic(self) -> None:
        candidates = [
            _candidate("a", "alpha", source="C:/docs/a.pdf", page=1),
            _candidate("b", "beta", source="C:/docs/b.pdf", page=2),
        ]
        reranker = MockReranker({"a": 0.1, "b": 0.8})

        first = reranker.rerank("question", candidates)
        second = reranker.rerank("question", candidates)

        self.assertEqual(
            [item["chunk_id"] for item in first.candidates],
            [item["chunk_id"] for item in second.candidates],
        )
        self.assertEqual(first.candidates[0]["reranker_score"], 0.8)

    def test_selector_records_score_diversity_redundancy_and_budget_discards(
        self,
    ) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=100)
        selector = EvidenceSelector(
            manager,
            strategies=(
                "top_k",
                "reranker_score_threshold",
                "source_diversity",
                "redundancy_reduction",
                "token_budget",
            ),
            max_chunks=3,
            score_threshold=0.5,
            token_budget=28,
            max_chunks_per_source=1,
            redundancy_threshold=0.8,
        )
        candidates = [
            _candidate("a", "alpha control", source="C:/docs/a.pdf", page=1)
            | {"reranker_score": 0.95},
            _candidate("b", "alpha control", source="C:/docs/b.pdf", page=2)
            | {"reranker_score": 0.90},
            _candidate("c", "different control", source="C:/docs/a.pdf", page=3)
            | {"reranker_score": 0.85},
            _candidate("d", "weak", source="C:/docs/d.pdf", page=4)
            | {"reranker_score": 0.1},
            _candidate(
                "e",
                "this evidence is too long",
                source="C:/docs/e.pdf",
                page=5,
            )
            | {"reranker_score": 0.8},
        ]

        result = selector.select(candidates)

        self.assertEqual([item["chunk_id"] for item in result.selected], ["a"])
        reasons = {item["chunk_id"]: item["discard_reason"] for item in result.discarded}
        self.assertEqual(reasons["b"], "redundancy")
        self.assertEqual(reasons["c"], "source_diversity_limit")
        self.assertEqual(reasons["d"], "threshold_failed")
        self.assertEqual(reasons["e"], "token_budget")
        self.assertLessEqual(result.selected_tokens, 28)

    def test_selector_respects_maximum_evidence_count(self) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=100)
        selector = EvidenceSelector(
            manager,
            strategies=("top_k",),
            max_chunks=2,
            score_threshold=0.0,
            token_budget=100,
            max_chunks_per_source=10,
            redundancy_threshold=1.0,
        )
        candidates = [
            _candidate(
                f"chunk-{index}",
                f"evidence {index}",
                source=f"C:/docs/{index}.pdf",
                page=index,
            )
            | {"reranker_score": 1.0 - index / 10}
            for index in range(1, 5)
        ]

        result = selector.select(candidates)

        self.assertEqual(len(result.selected), 2)
        self.assertTrue(
            all(
                item["discard_reason"] == "lower_rank_fallback"
                for item in result.discarded
            )
        )

    def test_selector_falls_back_to_floor_and_token_target(self) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=1_200)
        selector = EvidenceSelector(
            manager,
            strategies=(
                "top_k",
                "reranker_score_threshold",
                "source_diversity",
                "redundancy_reduction",
                "token_budget",
            ),
            min_selected_evidence=3,
            max_selected_evidence=8,
            score_threshold=0.2,
            token_budget=1_200,
            max_chunks_per_source=2,
            redundancy_threshold=0.9,
            fallback_to_top_n_if_empty=True,
            fallback_top_n=3,
            target_min_tokens=500,
            target_max_tokens=900,
        )
        candidates = [
            _candidate(
                f"weak-{index}",
                (f"distinct evidence section {index} " * 8)[:200],
                source=f"C:/docs/{index}.pdf",
                page=index,
            )
            | {"reranker_score": -float(index)}
            for index in range(1, 6)
        ]

        result = selector.select(candidates)

        self.assertGreaterEqual(len(result.selected), 3)
        self.assertGreaterEqual(result.selected_tokens, 500)
        self.assertLessEqual(result.selected_tokens, 900)
        self.assertTrue(result.fallback_used)
        self.assertTrue(result.weak_evidence)
        self.assertTrue(
            all(item["weak_evidence"] for item in result.selected)
        )
        self.assertTrue(
            all(
                item["discard_reason"]
                in {
                    "threshold_failed",
                    "redundancy",
                    "source_diversity_limit",
                    "token_budget",
                    "empty_text",
                    "lower_rank_fallback",
                }
                for item in result.discarded
            )
        )

    def test_selector_returns_zero_only_for_empty_candidates(self) -> None:
        manager = TokenBudgetManager(_CharacterTokenizer(), max_tokens=100)
        selector = EvidenceSelector(
            manager,
            strategies=("top_k", "reranker_score_threshold"),
            min_selected_evidence=3,
            max_selected_evidence=8,
            score_threshold=0.2,
            token_budget=100,
            max_chunks_per_source=2,
            redundancy_threshold=0.9,
            target_min_tokens=50,
            target_max_tokens=100,
        )
        candidates = [
            _candidate("empty", "", source="C:/docs/a.pdf", page=1)
            | {"reranker_score": 10.0}
        ]

        result = selector.select(candidates)

        self.assertEqual(result.selected, ())
        self.assertEqual(result.usable_candidate_count, 0)
        self.assertEqual(result.discarded[0]["discard_reason"], "empty_text")


class EvidenceQualityAndTraceTests(unittest.TestCase):
    def test_quality_scoring_reports_provenance_metadata_and_strength(self) -> None:
        scorer = EvidenceQualityScorer(
            strong_threshold=0.7,
            medium_threshold=0.4,
        )
        selected = [
            _candidate("a", "alpha", source="C:/docs/a.pdf", page=1)
            | {
                "reranker_score": 0.8,
                "retrieval_sources": ["dense", "bm25"],
                "evidence_token_count": 5,
            },
            _candidate("b", "beta", source="C:/docs/b.pdf", page=2)
            | {
                "reranker_score": 0.45,
                "retrieval_sources": ["dense"],
                "evidence_token_count": 4,
            },
        ]

        report = scorer.score(selected)

        self.assertEqual(report.chunks[0]["retrieval_source"], "both")
        self.assertEqual(report.chunks[0]["evidence_strength"], "strong")
        self.assertTrue(report.chunks[0]["metadata_complete"])
        self.assertEqual(report.summary["unique_source_count"], 2)
        self.assertEqual(
            report.summary["strength_distribution"],
            {"strong": 1, "medium": 1, "weak": 0},
        )

    def test_phase4_trace_round_trip(self) -> None:
        trace = Phase4Trace.from_dict(
            {
                "question": "What is the control?",
                "selected_chunks": [{"chunk_id": "a"}],
                "artifacts": {"context": Path("context/a.md")},
            }
        )

        restored = Phase4Trace.from_json(trace.to_json())

        self.assertEqual(restored.to_dict()["question"], "What is the control?")
        self.assertEqual(
            restored.to_dict()["artifacts"]["context"],
            str(Path("context/a.md")),
        )

    def test_starvation_diagnostics_flag_high_risk_selection(self) -> None:
        diagnostics = phase4_diagnostics(
            token_reduction_percent=95.0,
            average_reranker_score=0.0,
            medium_score_threshold=0.4,
            discarded=[],
            latency={},
            unique_source_count=0,
            selected_chunk_count=0,
            candidate_chunk_count=12,
            selected_evidence_tokens=0,
            answer_status="insufficient_evidence",
        )

        signals = {item["signal"] for item in diagnostics}
        self.assertIn("evidence_starvation", signals)
        self.assertIn("excessive_token_reduction", signals)
        self.assertIn("zero_average_reranker_score", signals)


class Phase4PipelineAndArtifactTests(unittest.TestCase):
    def _pipeline(self, root: Path) -> Phase4RAGPipeline:
        first = _candidate(
            "first",
            "The selected control requires review.",
            source=str(root / "manual.pdf"),
            page=1,
            score=0.8,
        )
        second = _candidate(
            "second",
            "Unrelated cafeteria information.",
            source=str(root / "other.pdf"),
            page=2,
            score=0.7,
        )
        config = Phase4Config(
            project_root=root,
            max_context_tokens=300,
            evidence_token_budget=120,
            max_query_variants=1,
            evidence_max_chunks=2,
            evidence_score_threshold=0.5,
            evidence_max_chunks_per_source=2,
        )
        return Phase4RAGPipeline(
            config,
            llm=_CitingLLM(),
            tokenizer=_CharacterTokenizer(),
            retrievers={
                "dense": _StaticRetriever("dense", [first, second]),
                "bm25": _StaticRetriever("bm25", [first, second]),
            },
            reranker=MockReranker({"first": 0.95, "second": 0.1}),
        )

    def test_pipeline_reranks_selects_and_preserves_phase3_keys(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            response = pipeline.answer("What control is required?")

        self.assertIn("context_stages", response)
        self.assertIn("token_usage", response)
        self.assertEqual(response["retrieval_mode"], "hybrid")
        self.assertEqual(
            [item["chunk_id"] for item in response["selected_evidence"]],
            ["first", "second"],
        )
        self.assertEqual(response["discarded_evidence"], [])
        self.assertEqual(
            response["selected_evidence"][1]["selection_reason"],
            "adaptive_fallback",
        )
        self.assertEqual(response["evidence_confidence"], "mixed")
        self.assertGreaterEqual(
            response["token_efficiency"]["token_reduction_percent"],
            0,
        )
        self.assertEqual(
            response["question_trace"]["pipeline_flow"][2],
            "reranker",
        )
        self.assertIn("Produce a detailed synthesis", pipeline.llm.prompt)
        self.assertIn(
            "Choose the answer structure that best fits the question",
            pipeline.llm.prompt,
        )
        self.assertIn("Do not use every section by default", pipeline.llm.prompt)
        self.assertIn("Question-shape guidance", pipeline.llm.prompt)
        self.assertIn("Step-by-Step Procedure", pipeline.llm.prompt)
        self.assertIn("Priority Matrix", pipeline.llm.prompt)
        self.assertIn("Cite every key factual claim", pipeline.llm.prompt)
        self.assertIn("Aim for at least 250 words", pipeline.llm.prompt)
        self.assertNotIn(
            "organize the answer as:\n- Executive answer",
            pipeline.llm.prompt,
        )
        self.assertNotIn("Answer concisely.", pipeline.llm.prompt)

    def test_fixed_answer_sections_preserve_previous_template(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.config.adaptive_answer_sections = False
            pipeline.answer("What control is required?")

        prompt = pipeline.llm.prompt
        self.assertIn(
            "organize the answer as:\n- Executive answer",
            prompt,
        )
        self.assertIn("- Evidence-backed findings", prompt)
        self.assertIn("- Operational implications", prompt)
        self.assertIn("- Recommended controls or actions", prompt)
        self.assertIn("- Risks, gaps, and caveats", prompt)
        self.assertIn("Include a short Decision notes section", prompt)
        self.assertNotIn("Question-shape guidance", prompt)

    def test_max_answer_words_is_added_to_phase4_prompt(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.config.max_answer_words = 450
            pipeline.answer("What control is required?")

        self.assertIn("Do not exceed 450 words", pipeline.llm.prompt)

    def test_phase3_prompt_style_remains_unchanged(self) -> None:
        prompt = build_grounded_prompt("Question?", "[1] Evidence.")

        self.assertIn("Answer concisely.", prompt)
        self.assertIn("Prefer 5", prompt)
        self.assertNotIn("Decision notes", prompt)

    def test_pipeline_answer_downloads_once_then_uses_cached_model(self) -> None:
        cache_state = {"available": False}
        local_only_calls: list[bool] = []

        def model_factory(*_: Any, **kwargs: Any) -> _FakeCrossEncoder:
            local_only = bool(kwargs["local_files_only"])
            local_only_calls.append(local_only)
            if local_only and not cache_state["available"]:
                raise OSError("cache miss")
            if not local_only:
                cache_state["available"] = True
            return _FakeCrossEncoder()

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first_pipeline = self._pipeline(root)
            first_pipeline.reranker = CrossEncoderReranker(
                first_pipeline.config.reranker_model_name,
                local_files_only=False,
            )
            second_pipeline = self._pipeline(root)
            second_pipeline.reranker = CrossEncoderReranker(
                second_pipeline.config.reranker_model_name,
                local_files_only=False,
            )
            output = io.StringIO()
            with patch(
                "sentence_transformers.CrossEncoder",
                side_effect=model_factory,
            ), redirect_stdout(output):
                first_response = first_pipeline.answer("What control is required?")
                second_response = second_pipeline.answer("What control is required?")

        self.assertEqual(local_only_calls, [True, False, True])
        self.assertEqual(first_pipeline.reranker.load_source, "download")
        self.assertEqual(second_pipeline.reranker.load_source, "cache")
        self.assertTrue(first_response["answer"])
        self.assertTrue(second_response["answer"])
        self.assertIn("downloaded and cached successfully", output.getvalue())
        self.assertIn("loaded from local Hugging Face cache", output.getvalue())

    def test_pipeline_answers_with_caution_when_only_weak_evidence_exists(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.reranker = MockReranker(
                {"first": -5.0, "second": -6.0}
            )
            response = pipeline.answer("What control is required?")

        self.assertEqual(response["answer_status"], "answered")
        self.assertEqual(response["evidence_confidence"], "weak")
        self.assertTrue(response["weak_evidence"])
        self.assertIn("Caution", response["answer"])
        self.assertGreater(len(response["selected_evidence"]), 0)
        self.assertIn(
            "All selected evidence is below the reranker threshold",
            pipeline.llm.prompt,
        )

    def test_generation_retry_succeeds_without_repeating_retrieval(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.config.generation_retries = 2
            pipeline.config.retry_cooldown_seconds = 0
            pipeline.llm = _FlakyLLM(failures=1)
            response = pipeline.answer("What control is required?")

        self.assertEqual(response["answer_status"], "answered")
        self.assertEqual(pipeline.llm.calls, 2)
        self.assertEqual(pipeline.metrics["generation_attempts"], 2.0)
        self.assertEqual(pipeline.metrics["generation_retry_count"], 1.0)
        self.assertIn("Recovered grounded answer", response["answer"])

    def test_exhausted_generation_is_exported_as_failed_row(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self._pipeline(root)
            pipeline.config.generation_retries = 1
            pipeline.config.retry_cooldown_seconds = 0
            pipeline.llm = _FlakyLLM(failures=99)
            result = Phase4Runner(
                pipeline=_ReadyPipeline(pipeline),
                config=pipeline.config,
            ).run(
                questions=["What control is required?"],
                run_mode="smoke",
            )

            with result.paths.results_csv.open(
                encoding="utf-8-sig",
                newline="",
            ) as handle:
                row = next(csv.DictReader(handle))
            checkpoint = json.loads(
                (result.paths.root / "checkpoint.json").read_text(
                    encoding="utf-8"
                )
            )
            report_exists = result.paths.report_html.is_file()
            metrics_exists = result.paths.metrics_json.is_file()

        self.assertEqual(row["answer_status"], "generation_failed")
        self.assertEqual(row["status"], "failed")
        self.assertIn("RuntimeError", row["error"])
        self.assertIn("std::bad_alloc", row["error"])
        self.assertEqual(len(checkpoint["failed_questions"]), 1)
        self.assertEqual(checkpoint["status"], "completed_with_failures")
        self.assertTrue(report_exists)
        self.assertTrue(metrics_exists)

    def test_generator_safe_failure_with_usable_evidence_becomes_grounded_fallback(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.llm = _RefusingLLM()
            response = pipeline.answer("What control is required?")

        self.assertEqual(response["answer_status"], "answered")
        self.assertGreater(len(response["selected_evidence"]), 0)
        self.assertIn("evidence review required", response["answer"].lower())
        self.assertTrue(response["citations"])
        self.assertTrue(response["extractive_fallback_used"])
        self.assertFalse(response["fallback_blocked"])

    def test_unsupported_current_question_is_not_answered(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            response = pipeline.answer(
                "What is today's airport weather forecast?"
            )

        self.assertEqual(response["answer_status"], "unsupported_query")
        self.assertEqual(response["raw_answer"], UNSUPPORTED_QUERY_RESPONSE)
        self.assertEqual(response["answer"], UNSUPPORTED_QUERY_RESPONSE)
        self.assertEqual(response["citations"], [])
        self.assertTrue(response["unsupported_query_detected"])

    def test_weak_evidence_does_not_become_extractive_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            pipeline.reranker = MockReranker(
                {"first": -5.0, "second": -6.0}
            )
            pipeline.llm = _RefusingLLM()
            response = pipeline.answer("What control is required?")

        self.assertEqual(
            response["answer_status"],
            "insufficient_evidence",
        )
        self.assertEqual(response["answer"], INSUFFICIENT_EVIDENCE_RESPONSE)
        self.assertEqual(response["citations"], [])
        self.assertFalse(response["extractive_fallback_used"])
        self.assertTrue(response["fallback_blocked"])

    def test_unsupported_status_is_preserved_in_csv_xlsx_and_html(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            pipeline = self._pipeline(Path(directory))
            result = Phase4Runner(
                pipeline=_ReadyPipeline(pipeline),
                config=pipeline.config,
            ).run(
                questions=["What is the current share price?"],
                run_mode="smoke",
            )
            with result.paths.results_csv.open(
                encoding="utf-8-sig",
                newline="",
            ) as handle:
                row = next(csv.DictReader(handle))
            workbook = load_workbook(result.paths.results_xlsx)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            xlsx_status = sheet.cell(
                row=2,
                column=headers.index("answer_status") + 1,
            ).value
            report = result.paths.report_html.read_text(encoding="utf-8")

        self.assertEqual(row["answer_status"], "Unsupported Query")
        self.assertEqual(xlsx_status, "Unsupported Query")
        self.assertIn("status-unsupported-query", report)
        self.assertEqual(result.metrics["unsupported_query_count"], 1)
        self.assertEqual(result.metrics["extractive_fallback_count"], 0)

    def test_insufficient_evidence_requires_no_usable_chunks(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config = Phase4Config(
                project_root=root,
                max_context_tokens=300,
                evidence_token_budget=120,
                max_query_variants=1,
            )
            pipeline = Phase4RAGPipeline(
                config,
                llm=_CitingLLM(),
                tokenizer=_CharacterTokenizer(),
                retrievers={
                    "dense": _StaticRetriever("dense", []),
                    "bm25": _StaticRetriever("bm25", []),
                },
                reranker=MockReranker({}),
            )
            response = pipeline.answer("What control is required?")

        self.assertEqual(response["answer_status"], "insufficient_evidence")
        self.assertEqual(response["selected_evidence"], [])
        self.assertEqual(response["citations"], [])

    def test_phase4_bundle_contains_all_artifacts_and_report_sections(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            (root / "other.pdf").write_bytes(b"%PDF-test")
            pipeline = self._pipeline(root)
            pipeline.config.min_selected_evidence = 1
            pipeline.config.max_selected_evidence = 1
            pipeline.config.selected_evidence_target_min_tokens = 1
            pipeline.config.selected_evidence_target_max_tokens = 120
            pipeline.on_config_changed()
            result = Phase4Runner(
                pipeline=_ReadyPipeline(pipeline),
                config=pipeline.config,
            ).run(
                questions=["What control is required?"],
                run_mode="smoke",
            )

            paths = result.paths
            required = (
                paths.results_csv,
                paths.results_xlsx,
                paths.report_html,
                paths.config_json,
                paths.summary_json,
                paths.metrics_json,
                paths.retrieval_json,
                paths.logs,
                paths.root / "file_format_summary.csv",
                paths.root / "file_extension_distribution.csv",
                paths.root / "skipped_files.csv",
            )
            self.assertTrue(all(path.is_file() for path in required))
            self.assertEqual(len(list(paths.context.glob("*.md"))), 1)
            self.assertGreaterEqual(len(list(paths.figures.glob("*.svg"))), 8)

            header = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[0].split(",")
            expected_columns = (
                CSV_COLUMNS
                + PHASE2_CSV_COLUMNS
                + PHASE3_CSV_COLUMNS
                + PHASE4_CSV_COLUMNS
                + ["run_mode"]
            )
            self.assertEqual(header, expected_columns)
            self.assertEqual(
                header[: len(CSV_COLUMNS + PHASE2_CSV_COLUMNS + PHASE3_CSV_COLUMNS)],
                CSV_COLUMNS + PHASE2_CSV_COLUMNS + PHASE3_CSV_COLUMNS,
            )
            for column in PHASE4_CSV_COLUMNS:
                self.assertIn(column, header)
            workbook = load_workbook(paths.results_xlsx)
            self.assertIn("file_format_summary", workbook.sheetnames)
            self.assertIn("file_extension_distribution", workbook.sheetnames)
            self.assertIn("skipped_files", workbook.sheetnames)
            workbook_header = [
                cell.value
                for cell in next(workbook.active.iter_rows(min_row=1, max_row=1))
            ]
            self.assertEqual(workbook_header, expected_columns)
            pdf_column = header.index("pdf_links") + 1
            self.assertIsNotNone(workbook.active.cell(2, pdf_column).hyperlink)

            report = paths.report_html.read_text(encoding="utf-8")
            for heading in (
                "Executive Summary",
                "Answers",
                "Citations",
                "Reranking Trace",
                "Evidence Selection",
                "Token Reduction",
                "Latency Breakdown",
                "Evidence Quality",
                "Enterprise File Format Readiness",
                "OCR Processing Summary",
                "Source Diversity",
                "Selected vs Discarded Chunks",
                "Discard Reason Breakdown",
                "Phase 3 vs Phase 4 Comparison",
            ):
                self.assertIn(heading, report)
            self.assertNotIn("https://cdn", report)
            self.assertNotIn("<script src=", report)
            self.assertIn("<strong>Use the selected control</strong>", report)
            self.assertIn(
                "Full generated answers are rendered below without preview truncation.",
                report,
            )
            self.assertIn("Evidence selection reduces irrelevant context", report)

            trace = json.loads(
                paths.retrieval_json.read_text(encoding="utf-8")
            )[0]
            self.assertEqual(trace["selected_chunks"][0]["chunk_id"], "first")
            self.assertEqual(
                trace["discarded_chunks"][0]["discard_reason"],
                "threshold_failed",
            )
            self.assertIn("results_csv", trace["artifacts"])
            metrics = json.loads(paths.metrics_json.read_text(encoding="utf-8"))
            self.assertEqual(
                metrics["discard_reason_distribution"]["threshold_failed"],
                1,
            )

    def test_previous_phase_defaults_remain_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            phase1 = KnowledgeOSConfig(project_root=root)
            phase2 = Phase2Config(project_root=root)
            phase3 = Phase3Config(project_root=root)
            phase4 = Phase4Config(project_root=root)

        self.assertEqual(phase1.qdrant_collection_name, "cial_basic_rag")
        self.assertEqual(phase2.qdrant_collection_name, "cial_phase2")
        self.assertEqual(phase3.qdrant_collection_name, "cial_phase3")
        self.assertEqual(phase3.phase_output_name, "03_Hybrid_Retrieval")
        self.assertEqual(phase4.phase_output_name, "04_Reranking_and_Evidence_Selection")
        self.assertFalse(phase4.enable_neighbor_expansion)
        self.assertFalse(phase4.reranker_local_files_only)
        self.assertEqual(phase4.min_selected_evidence, 3)
        self.assertEqual(phase4.max_selected_evidence, 8)
        self.assertEqual(phase4.reranker_score_threshold, -4.0)
        self.assertTrue(phase4.fallback_to_top_n_if_empty)
        self.assertEqual(phase4.fallback_top_n, 3)
        self.assertTrue(phase4.weak_evidence_answer_allowed)
        self.assertEqual(phase4.answer_detail_level, "detailed")
        self.assertEqual(phase4.min_answer_words, 250)
        self.assertIsNone(phase4.max_answer_words)
        self.assertTrue(phase4.prefer_structured_answers)
        self.assertTrue(phase4.adaptive_answer_sections)
        self.assertTrue(phase4.include_decision_notes)
        self.assertEqual(phase4.generation_retries, 2)
        self.assertEqual(phase4.retry_cooldown_seconds, 20.0)


if __name__ == "__main__":
    unittest.main()
