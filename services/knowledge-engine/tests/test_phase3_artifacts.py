from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cial_knowledge_os.batch_qa import CSV_COLUMNS, PHASE2_CSV_COLUMNS, PHASE3_CSV_COLUMNS
from cial_knowledge_os.config import Phase3Config
from cial_knowledge_os.phase3_runner import Phase3Runner
from cial_knowledge_os.run_manager import RunManager


class _ArtifactPipeline:
    def __init__(self, config: Phase3Config) -> None:
        self.config = config
        self.metrics = {
            "retrieval_latency": 0.01,
            "generation_latency": 0.02,
        }

    @property
    def is_ready_for_answering(self) -> bool:
        return True

    def answer(self, question: str):
        evidence = {
            "text": "Grounded evidence.",
            "score": 0.03,
            "rrf_score": 0.03,
            "retrieval_sources": ["dense", "bm25"],
            "source": "manual.pdf",
            "page_number": 4,
            "chunk_id": "manual:p4:c1",
            "metadata": {
                "source": str(self.config.project_root / "manual.pdf"),
                "file_name": "manual.pdf",
                "page_number": 4,
                "chunk_id": "manual:p4:c1",
                "chunk_index": 1,
            },
        }
        link = (self.config.project_root / "manual.pdf").resolve().as_uri() + "#page=4"
        return {
            "question": question,
            "answer": "Grounded answer [1].",
            "raw_answer": "Grounded answer [1].",
            "answer_status": "answered",
            "retrieval_mode": "hybrid",
            "retrieved": [evidence],
            "query_variants": [{"technique": "original", "query": question}],
            "retrieved_by_query": {"original": [evidence]},
            "context_stages": {
                "retrieved": [evidence],
                "deduplicated": [evidence],
                "expanded": [evidence],
                "merged": [evidence],
                "compressed": [evidence],
            },
            "stage_counts": {
                "retrieved": 1,
                "deduplicated": 1,
                "expanded": 1,
                "merged": 1,
                "compressed": 1,
            },
            "context": "Grounded evidence.",
            "prompt": "Grounded prompt.",
            "token_usage": {
                "budget": 100,
                "used": 20,
                "remaining": 80,
                "context_tokens": 20,
                "encoding_name": "cl100k_base",
                "truncated_sections": 0,
                "omitted_sections": 0,
                "budget_type": "tokens",
            },
            "citations": [
                {
                    "reference_id": 1,
                    "source": "manual.pdf",
                    "source_file": "manual.pdf",
                    "source_path": str(self.config.project_root / "manual.pdf"),
                    "page_number": 4,
                    "chunk_id": "manual:p4:c1",
                    "pdf_link": link,
                }
            ],
        }


class _MarkdownArtifactPipeline(_ArtifactPipeline):
    def answer(self, question: str):
        response = super().answer(question)
        response["answer"] = (
            "**Control summary**\n\n"
            "* Enable filtering\n"
            "* Track `AIBOM`\n\n"
            "1. Review evidence\n"
            "2. Record findings\n\n"
            "<script>alert('unsafe')</script>\n\n"
            "References:\n"
            "[1] manual.pdf | page 4 | chunk manual:p4:c1"
        )
        response["raw_answer"] = response["answer"]
        response["context_stages"]["compressed"][0]["text"] = (
            "Grounded <unsafe> evidence."
        )
        response["citations"][0]["score"] = 0.0325
        return response


class _TraceArtifactPipeline(_ArtifactPipeline):
    def answer(self, question: str):
        response = super().answer(question)
        citation = response["citations"][0] | {
            "score": 0.0325,
            "retrieval_sources": ["dense", "bm25"],
            "retrieval_source": "both",
        }
        result = {
            "rank": 1,
            "query_variant": "original",
            "source": "manual.pdf",
            "page": 4,
            "chunk_id": "manual:p4:c1",
            "score": 0.0325,
            "rrf_score": 0.0325,
            "dense_rank": 1,
            "bm25_rank": 1,
            "retrieval_sources": ["dense", "bm25"],
            "retrieval_source": "both",
            "matched_terms": ["control"],
            "text_preview": "Grounded evidence.",
            "token_count": 4,
            "citation_link": citation["pdf_link"],
        }
        response["question_trace"] = {
            "question": question,
            "retrieval_mode": "hybrid",
            "query_variants": [{"technique": "original", "query": question}],
            "dense_results": [result],
            "bm25_results": [result],
            "fused_results": [result],
            "overlap": {
                "dense_only_count": 0,
                "bm25_only_count": 0,
                "both_count": 1,
                "union_count": 1,
            },
            "deduplication": {
                "before": 1,
                "after": 1,
                "duplicates_removed": 0,
                "key": "source + page + chunk_id",
            },
            "neighbor_expansion": {
                "original_chunks": 1,
                "neighbors_added": 0,
                "total_after_expansion": 1,
            },
            "context_funnel": {
                "counts": {
                    "dense_raw": 1,
                    "bm25_raw": 1,
                    "combined": 2,
                    "fused": 1,
                    "retrieved": 1,
                    "deduplicated": 1,
                    "expanded": 1,
                    "merged": 1,
                    "compressed": 1,
                },
                "token_counts": {
                    "retrieved": 4,
                    "deduplicated": 4,
                    "expanded": 4,
                    "merged": 4,
                    "compressed": 4,
                    "final_context": 4,
                },
            },
            "token_usage": {
                **response["token_usage"],
                "max_context_tokens": 100,
                "context_tokens_used": 20,
                "remaining_tokens": 80,
                "chunks_included": 1,
                "chunks_skipped": 0,
                "utilization_percent": 20.0,
            },
            "final_context": "Grounded evidence.",
            "final_context_chunks": [result],
            "generation": {
                "model_name": "test-model",
                "prompt_tokens": 10,
                "context_tokens": 4,
                "answer_tokens": 5,
                "latency_seconds": 0.02,
                "status": "answered",
            },
            "answer": response["raw_answer"],
            "citations": [citation],
            "latency": {
                "retrieval_seconds": 0.01,
                "context_construction_seconds": 0.005,
                "generation_seconds": 0.02,
                "total_pipeline_seconds": 0.035,
                "artifact_export_seconds": None,
            },
            "source_diversity": {
                "unique_source_count": 1,
                "unique_page_count": 1,
                "sources": ["manual.pdf"],
            },
            "decision_summary": [
                {
                    "signal": "bm25_value",
                    "recommendation": "BM25 agreed with dense retrieval.",
                }
            ],
            "artifacts": {},
        }
        return response


class RunManagerTests(unittest.TestCase):
    def test_run_directories_never_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            config = Phase3Config(project_root=Path(directory))
            first = RunManager.from_config(config).create()
            second = RunManager.from_config(config).create()
        self.assertNotEqual(first.root, second.root)


class Phase3ArtifactTests(unittest.TestCase):
    def test_complete_standalone_bundle_is_generated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_ArtifactPipeline(config),
                config=config,
            ).run(questions=["What is the control?"])
            paths = result.paths

            required_files = (
                paths.results_csv,
                paths.results_xlsx,
                paths.report_html,
                paths.config_json,
                paths.summary_json,
                paths.retrieval_json,
                paths.metrics_json,
                paths.logs,
            )
            self.assertTrue(all(path.is_file() for path in required_files))
            self.assertTrue(paths.figures.is_dir())
            self.assertTrue(
                (paths.figures / config.artifact_names.latency_figure).is_file()
            )
            self.assertEqual(len(list(paths.context.glob("*.md"))), 1)

            csv_header = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[0].split(",")
            self.assertEqual(
                csv_header,
                [*CSV_COLUMNS, *PHASE2_CSV_COLUMNS, *PHASE3_CSV_COLUMNS],
            )
            csv_values = paths.results_csv.read_text(
                encoding="utf-8-sig"
            ).splitlines()[1]
            self.assertIn("cl100k_base", csv_values)
            workbook = load_workbook(paths.results_xlsx)
            sheet = workbook.active
            pdf_column = csv_header.index("pdf_links") + 1
            self.assertIsNotNone(sheet.cell(2, pdf_column).hyperlink)

            report = paths.report_html.read_text(encoding="utf-8")
            self.assertIn("Executive Summary", report)
            self.assertIn("Retrieved Context", report)
            self.assertIn("Token Usage", report)
            self.assertNotIn("https://cdn", report)
            self.assertNotIn("<script src=", report)
            self.assertEqual(
                json.loads(paths.summary_json.read_text(encoding="utf-8"))[
                    "question_count"
                ],
                1,
            )
            self.assertEqual(
                json.loads(paths.config_json.read_text(encoding="utf-8"))[
                    "tokenizer_encoding_name"
                ],
                "cl100k_base",
            )
            self.assertIn(
                '"event":"run"',
                paths.logs.read_text(encoding="utf-8"),
            )

    def test_html_renders_safe_markdown_and_structured_citations(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_MarkdownArtifactPipeline(config),
                config=config,
            ).run(questions=["What is the control?"])

            csv_text = result.paths.results_csv.read_text(encoding="utf-8-sig")
            self.assertIn("**Control summary**", csv_text)
            workbook = load_workbook(result.paths.results_xlsx)
            answer_column = CSV_COLUMNS.index("answer") + 1
            self.assertIn(
                "**Control summary**",
                workbook.active.cell(2, answer_column).value,
            )

            report = result.paths.report_html.read_text(encoding="utf-8")
            answer_start = report.index('<div class="answer-content">')
            answer_end = report.index("</div>", answer_start)
            answer_html = report[answer_start:answer_end]
            self.assertIn("<strong>Control summary</strong>", answer_html)
            self.assertIn("<ul><li>Enable filtering</li>", answer_html)
            self.assertIn("<ol><li>Review evidence</li>", answer_html)
            self.assertIn("<code>AIBOM</code>", answer_html)
            self.assertNotIn("**Control summary**", answer_html)
            self.assertNotIn("<pre>", answer_html)
            self.assertNotIn("<script>", answer_html)
            self.assertIn("&lt;script&gt;", answer_html)
            self.assertNotIn("References:", answer_html)

            self.assertIn('class="citation-list"', report)
            self.assertIn('class="citation-card"', report)
            self.assertIn("manual.pdf", report)
            self.assertIn("Page 4", report)
            self.assertIn("Chunk manual:p4:c1", report)
            self.assertIn("Score 0.0325", report)
            self.assertIn("Open PDF", report)
            self.assertIn("file:///", report)
            self.assertIn(
                "<pre>Grounded &lt;unsafe&gt; evidence.</pre>",
                report,
            )

    def test_optional_run_metadata_is_exported_without_changing_default_schema(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_ArtifactPipeline(config),
                config=config,
            ).run(
                questions=["What is the control?"],
                run_metadata={
                    "run_type": "manual_qa",
                    "run_label": "interactive_manual_qa",
                },
            )

            csv_text = result.paths.results_csv.read_text(encoding="utf-8-sig")
            header = csv_text.splitlines()[0].split(",")
            self.assertEqual(header[-2:], ["run_type", "run_label"])
            self.assertIn("interactive_manual_qa", csv_text)

            config_payload = json.loads(
                result.paths.config_json.read_text(encoding="utf-8")
            )
            self.assertEqual(
                config_payload["run_overrides"]["run_type"],
                "manual_qa",
            )
            summary = json.loads(
                result.paths.summary_json.read_text(encoding="utf-8")
            )
            self.assertEqual(summary["run_label"], "interactive_manual_qa")
            retrieval = json.loads(
                result.paths.retrieval_json.read_text(encoding="utf-8")
            )
            self.assertEqual(
                retrieval[0]["run_metadata"]["run_type"],
                "manual_qa",
            )
            report = result.paths.report_html.read_text(encoding="utf-8")
            self.assertIn("Run Label", report)
            self.assertIn("interactive_manual_qa", report)

    def test_question_trace_is_exported_to_all_phase3_report_surfaces(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "manual.pdf").write_bytes(b"%PDF-test")
            config = Phase3Config(project_root=root)
            result = Phase3Runner(
                pipeline=_TraceArtifactPipeline(config),
                config=config,
            ).run(questions=["What is the control?"])

            csv_text = result.paths.results_csv.read_text(encoding="utf-8-sig")
            header = csv_text.splitlines()[0].split(",")
            for name in (
                "dense_result_count",
                "bm25_result_count",
                "fused_result_count",
                "final_context_chunk_count",
                "context_tokens_used",
                "token_utilization",
                "generation_latency_seconds",
                "citation_count",
                "unique_source_count",
            ):
                self.assertIn(name, header)

            trace = json.loads(
                result.paths.retrieval_json.read_text(encoding="utf-8")
            )[0]
            self.assertEqual(trace["dense_results"][0]["dense_rank"], 1)
            self.assertEqual(trace["bm25_results"][0]["bm25_rank"], 1)
            self.assertEqual(trace["fused_results"][0]["rrf_score"], 0.0325)
            self.assertEqual(trace["deduplication"]["duplicates_removed"], 0)
            self.assertEqual(trace["token_usage"]["utilization_percent"], 20.0)
            self.assertIn("results_csv", trace["artifacts"])
            self.assertIsNotNone(
                trace["latency"]["artifact_export_seconds"]
            )

            report = result.paths.report_html.read_text(encoding="utf-8")
            for heading in (
                "Question Answers",
                "Per-Question Trace",
                "Retrieval Diagnostics",
                "Token Usage",
                "Latency Breakdown",
                "Citation Evidence",
            ):
                self.assertIn(heading, report)
            self.assertIn("RRF fusion", report)
            self.assertIn("Dense retrieval", report)
            self.assertIn("BM25 retrieval", report)


if __name__ == "__main__":
    unittest.main()
