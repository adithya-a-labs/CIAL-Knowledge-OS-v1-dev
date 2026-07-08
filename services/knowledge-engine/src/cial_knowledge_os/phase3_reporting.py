"""CSV, XLSX, and standalone HTML reports for one Phase 3 run."""

from __future__ import annotations

import csv
import html
import json
import logging
import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def write_results_csv(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> Path:
    """Write the established UTF-8-with-BOM batch CSV schema."""

    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns))
        writer.writeheader()
        writer.writerows(rows)
    logger.info(
        "csv_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _json_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        parsed = json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return []
    return parsed if isinstance(parsed, list) else []


def write_results_xlsx(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> Path:
    """Write a formatted workbook with clickable first-citation PDF links."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Install the pinned project "
            "dependencies from requirements.txt."
        ) from exc

    target = Path(path).expanduser().resolve()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Phase 3 Results"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for column_index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=name)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pdf_column = columns.index("pdf_links") + 1 if "pdf_links" in columns else None
    for row_index, row in enumerate(rows, start=2):
        for column_index, name in enumerate(columns, start=1):
            value = row.get(name, "")
            if isinstance(value, (dict, list, tuple)):
                value = json.dumps(value, ensure_ascii=False)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if pdf_column is not None:
            links = _json_list(row.get("pdf_links"))
            if links:
                cell = sheet.cell(row=row_index, column=pdf_column)
                cell.value = "Open first cited PDF"
                cell.hyperlink = str(links[0])
                cell.style = "Hyperlink"

    width_overrides = {
        "question": 42,
        "answer": 72,
        "retrieval_trace": 55,
        "error": 42,
        "pdf_links": 28,
    }
    for index, name in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            width_overrides.get(name, min(max(len(name) + 2, 14), 28))
        )
    sheet.row_dimensions[1].height = 28
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    logger.info(
        "xlsx_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _metric_card(label: str, value: Any) -> str:
    return (
        '<div class="metric"><span>'
        + html.escape(label)
        + "</span><strong>"
        + html.escape(str(value))
        + "</strong></div>"
    )


def write_latency_svg(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
) -> Path:
    """Write a dependency-free latency chart for the run figures directory."""

    values = [float(row.get("total_latency_seconds") or 0.0) for row in rows]
    width, height, margin = 900, 320, 45
    chart_width = width - 2 * margin
    chart_height = height - 2 * margin
    maximum = max(values, default=0.0) or 1.0
    bar_width = chart_width / max(len(values), 1)
    bars = []
    for index, value in enumerate(values):
        bar_height = chart_height * value / maximum
        x = margin + index * bar_width + 2
        y = margin + chart_height - bar_height
        bars.append(
            f'<rect x="{x:.2f}" y="{y:.2f}" width="{max(1.0, bar_width - 4):.2f}" '
            f'height="{bar_height:.2f}" fill="#1f6f8b"><title>Q{index + 1}: '
            f"{value:.4f}s</title></rect>"
        )
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}" role="img" aria-label="Question latency">'
        '<rect width="100%" height="100%" fill="#ffffff"/>'
        f'<line x1="{margin}" y1="{margin + chart_height}" x2="{width - margin}" '
        f'y2="{margin + chart_height}" stroke="#60717a"/>'
        f'<text x="{margin}" y="24" font-family="sans-serif" font-size="18" '
        f'fill="#12344d">Question latency (maximum {maximum:.4f}s)</text>'
        + "".join(bars)
        + "</svg>"
    )
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(svg, encoding="utf-8")
    logger.info(
        "latency_figure_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target


def _render_inline_markdown(value: str) -> str:
    """Render a deliberately small, HTML-escaped inline Markdown subset."""

    rendered: list[str] = []
    for part in re.split(r"(`[^`\n]*`)", value):
        if part.startswith("`") and part.endswith("`") and len(part) >= 2:
            rendered.append(f"<code>{html.escape(part[1:-1])}</code>")
            continue
        escaped = html.escape(part)
        escaped = re.sub(
            r"\*\*(.+?)\*\*",
            r"<strong>\1</strong>",
            escaped,
        )
        escaped = re.sub(
            r"__(.+?)__",
            r"<strong>\1</strong>",
            escaped,
        )
        escaped = re.sub(
            r"(?<!\*)\*([^*\n]+)\*(?!\*)",
            r"<em>\1</em>",
            escaped,
        )
        rendered.append(escaped)
    return "".join(rendered)


def _markdown_table_cells(value: str) -> list[str]:
    """Split one pipe-table row without allowing raw HTML through."""

    stripped = value.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [cell.strip() for cell in stripped.split("|")]


def _is_markdown_table_separator(value: str) -> bool:
    cells = _markdown_table_cells(value)
    return bool(cells) and all(
        re.fullmatch(r":?-{3,}:?", cell) is not None for cell in cells
    )


def _render_answer_markdown(value: str) -> str:
    """Safely render common enterprise Markdown without external dependencies."""

    output: list[str] = []
    paragraph: list[str] = []
    list_kind: str | None = None
    list_items: list[str] = []
    fenced_code: list[str] | None = None

    def flush_paragraph() -> None:
        if paragraph:
            output.append(
                "<p>"
                + "<br>".join(
                    _render_inline_markdown(line) for line in paragraph
                )
                + "</p>"
            )
            paragraph.clear()

    def flush_list() -> None:
        nonlocal list_kind
        if list_kind is not None:
            output.append(
                f"<{list_kind}>"
                + "".join(
                    f"<li>{_render_inline_markdown(item)}</li>"
                    for item in list_items
                )
                + f"</{list_kind}>"
            )
            list_kind = None
            list_items.clear()

    lines = str(value).splitlines()
    line_index = 0
    while line_index < len(lines):
        raw_line = lines[line_index]
        stripped = raw_line.strip()
        if stripped.startswith("```"):
            flush_paragraph()
            flush_list()
            if fenced_code is None:
                fenced_code = []
            else:
                output.append(
                    "<pre><code>"
                    + html.escape("\n".join(fenced_code))
                    + "</code></pre>"
                )
                fenced_code = None
            line_index += 1
            continue
        if fenced_code is not None:
            fenced_code.append(raw_line)
            line_index += 1
            continue
        if not stripped:
            flush_paragraph()
            flush_list()
            line_index += 1
            continue
        if (
            "|" in stripped
            and line_index + 1 < len(lines)
            and _is_markdown_table_separator(lines[line_index + 1])
        ):
            flush_paragraph()
            flush_list()
            headers = _markdown_table_cells(stripped)
            rows: list[list[str]] = []
            line_index += 2
            while line_index < len(lines):
                candidate = lines[line_index].strip()
                if not candidate or "|" not in candidate:
                    break
                rows.append(_markdown_table_cells(candidate))
                line_index += 1
            width = len(headers)
            output.append(
                '<div class="markdown-table-wrap"><table class="markdown-table">'
                "<thead><tr>"
                + "".join(
                    f"<th>{_render_inline_markdown(cell)}</th>"
                    for cell in headers
                )
                + "</tr></thead><tbody>"
                + "".join(
                    "<tr>"
                    + "".join(
                        f"<td>{_render_inline_markdown(cell)}</td>"
                        for cell in [*row[:width], *([""] * max(0, width - len(row)))]
                    )
                    + "</tr>"
                    for row in rows
                )
                + "</tbody></table></div>"
            )
            continue

        heading = re.match(r"^(#{1,6})\s+(.+)$", stripped)
        bullet = re.match(r"^[-*+]\s+(.+)$", stripped)
        numbered = re.match(r"^\d+[.)]\s+(.+)$", stripped)
        if heading:
            flush_paragraph()
            flush_list()
            level = len(heading.group(1))
            output.append(
                f"<h{level}>{_render_inline_markdown(heading.group(2))}</h{level}>"
            )
        elif bullet:
            flush_paragraph()
            if list_kind not in {None, "ul"}:
                flush_list()
            list_kind = "ul"
            list_items.append(bullet.group(1))
        elif numbered:
            flush_paragraph()
            if list_kind not in {None, "ol"}:
                flush_list()
            list_kind = "ol"
            list_items.append(numbered.group(1))
        else:
            flush_list()
            paragraph.append(stripped)
        line_index += 1

    if fenced_code is not None:
        output.append(
            "<pre><code>"
            + html.escape("\n".join(fenced_code))
            + "</code></pre>"
        )
    flush_paragraph()
    flush_list()
    return "".join(output)


def render_safe_markdown(value: str) -> str:
    """Public safe Markdown renderer shared by HTML and notebook trace views."""

    return _render_answer_markdown(value)


def _answer_without_reference_tail(
    answer: str,
    citations: Sequence[Mapping[str, Any]],
) -> str:
    """Remove the plain-text reference appendix when structured citations exist."""

    if not citations:
        return answer
    marker = re.search(r"(?im)^\s*references?\s*:\s*$", answer)
    return answer[: marker.start()].rstrip() if marker else answer


def _citation_html(citation: Mapping[str, Any]) -> str:
    source = citation.get("source_file") or citation.get("source") or "Unknown"
    reference_id = citation.get("reference_id", "?")
    parts: list[str] = []
    if citation.get("page_number") not in {None, ""}:
        parts.append(f"Page {html.escape(str(citation['page_number']))}")
    if citation.get("chunk_id") not in {None, ""}:
        parts.append(f"Chunk {html.escape(str(citation['chunk_id']))}")
    score = citation.get("score")
    if score not in {None, ""}:
        try:
            score_label = f"{float(score):.4f}"
        except (TypeError, ValueError):
            score_label = html.escape(str(score))
        parts.append(f"Score {score_label}")
    metadata = " — ".join(parts)
    link = citation.get("pdf_link")
    action = (
        f'<a href="{html.escape(str(link), quote=True)}">Open PDF</a>'
        if link
        else '<span class="no-link">PDF link unavailable</span>'
    )
    return (
        '<li class="citation-card">'
        f'<span class="citation-reference">[{html.escape(str(reference_id))}]</span>'
        '<div>'
        f'<strong>{html.escape(str(source))}</strong>'
        + (
            f'<div class="source-meta">{metadata}</div>'
            if metadata
            else ""
        )
        + f'<div class="citation-action">{action}</div>'
        "</div></li>"
    )


def _table_html(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> str:
    if not rows:
        return "<p>No data recorded.</p>"
    head = "".join(f"<th>{html.escape(value)}</th>" for value in headers)
    body = "".join(
        "<tr>"
        + "".join(f"<td>{html.escape(str(value))}</td>" for value in row)
        + "</tr>"
        for row in rows
    )
    return f'<div class="table-wrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>'


def _trace_results_html(
    results: Sequence[Mapping[str, Any]],
    *,
    fused: bool = False,
) -> str:
    headers = (
        [
            "Final rank",
            "Source",
            "Page",
            "Chunk",
            "Dense rank",
            "BM25 rank",
            "RRF score",
            "Contribution",
        ]
        if fused
        else [
            "Rank",
            "Variant",
            "Source",
            "Page",
            "Chunk",
            "Score",
            "Matched terms",
            "Preview",
        ]
    )
    rows: list[list[Any]] = []
    for result in results:
        if fused:
            rows.append(
                [
                    result.get("rank", ""),
                    result.get("source", ""),
                    result.get("page", ""),
                    result.get("chunk_id", ""),
                    result.get("dense_rank", ""),
                    result.get("bm25_rank", ""),
                    result.get("rrf_score", ""),
                    result.get("retrieval_source", ""),
                ]
            )
        else:
            rows.append(
                [
                    result.get("rank", ""),
                    result.get("query_variant", ""),
                    result.get("source", ""),
                    result.get("page", ""),
                    result.get("chunk_id", ""),
                    result.get("score", ""),
                    ", ".join(result.get("matched_terms") or []),
                    result.get("text_preview", ""),
                ]
            )
    return _table_html(headers, rows)


def _question_trace_html(index: int, trace: Mapping[str, Any]) -> str:
    variants = [
        value
        for value in (trace.get("query_variants") or [])
        if isinstance(value, Mapping)
    ]
    dense = [
        value
        for value in (trace.get("dense_results") or [])
        if isinstance(value, Mapping)
    ]
    bm25 = [
        value
        for value in (trace.get("bm25_results") or [])
        if isinstance(value, Mapping)
    ]
    fused = [
        value
        for value in (trace.get("fused_results") or [])
        if isinstance(value, Mapping)
    ]
    overlap = trace.get("overlap")
    overlap = overlap if isinstance(overlap, Mapping) else {}
    dedup = trace.get("deduplication")
    dedup = dedup if isinstance(dedup, Mapping) else {}
    neighbors = trace.get("neighbor_expansion")
    neighbors = neighbors if isinstance(neighbors, Mapping) else {}
    expanded_chunks = [
        value
        for value in (neighbors.get("expanded_chunks") or [])
        if isinstance(value, Mapping)
    ]
    funnel = trace.get("context_funnel")
    funnel = funnel if isinstance(funnel, Mapping) else {}
    counts = funnel.get("counts")
    counts = counts if isinstance(counts, Mapping) else {}
    token_counts = funnel.get("token_counts")
    token_counts = token_counts if isinstance(token_counts, Mapping) else {}
    final_chunks = [
        value
        for value in (trace.get("final_context_chunks") or [])
        if isinstance(value, Mapping)
    ]
    decisions = [
        value
        for value in (trace.get("decision_summary") or [])
        if isinstance(value, Mapping)
    ]
    variants_html = _table_html(
        ["Technique", "Query"],
        [
            [variant.get("technique", ""), variant.get("query", "")]
            for variant in variants
        ],
    )
    funnel_order = (
        "dense_raw",
        "bm25_raw",
        "combined",
        "fused",
        "retrieved",
        "deduplicated",
        "expanded",
        "merged",
        "compressed",
    )
    funnel_html = _table_html(
        ["Stage", "Chunks", "Tokens"],
        [
            [
                stage.replace("_", " ").title(),
                counts.get(stage, ""),
                token_counts.get(stage, ""),
            ]
            for stage in funnel_order
            if stage in counts
        ],
    )
    final_context_html = "".join(
        "<details><summary>"
        + html.escape(
            f"{chunk.get('source', 'Unknown')} — page {chunk.get('page', '')} "
            f"— chunk {chunk.get('chunk_id', '')} — "
            f"{chunk.get('retrieval_source', 'unknown')} — "
            f"{chunk.get('token_count', 0)} tokens"
        )
        + "</summary><pre>"
        + html.escape(str(chunk.get("text_preview") or ""))
        + "</pre></details>"
        for chunk in final_chunks
    )
    decisions_html = "".join(
        '<li class="diagnostic-card"><strong>'
        + html.escape(str(item.get("signal") or "diagnostic"))
        + "</strong><span>"
        + html.escape(str(item.get("recommendation") or ""))
        + "</span></li>"
        for item in decisions
    )
    return f"""<article class="trace-article">
<h3>Q{index}: {html.escape(str(trace.get("question") or ""))}</h3>
<details open><summary>Query transformations</summary>{variants_html}</details>
<details><summary>Dense retrieval ({len(dense)} results)</summary>{_trace_results_html(dense)}</details>
<details><summary>BM25 retrieval ({len(bm25)} results)</summary>{_trace_results_html(bm25)}</details>
<details><summary>RRF fusion ({len(fused)} results)</summary>{_trace_results_html(fused, fused=True)}</details>
<div class="grid">
{_metric_card("Dense only", overlap.get("dense_only_count", 0))}
{_metric_card("BM25 only", overlap.get("bm25_only_count", 0))}
{_metric_card("Both", overlap.get("both_count", 0))}
{_metric_card("Duplicates removed", dedup.get("duplicates_removed", 0))}
{_metric_card("Neighbors added", neighbors.get("neighbors_added", 0))}
</div>
<details><summary>Neighbor expansion ({len(expanded_chunks)} total chunks)</summary>
{_table_html(
    ["Source", "Page", "Chunk", "Added neighbor"],
    [
        [
            chunk.get("source", ""),
            chunk.get("page", ""),
            chunk.get("chunk_id", ""),
            chunk.get("is_neighbor", False),
        ]
        for chunk in expanded_chunks
    ],
)}</details>
<h4>Context construction funnel</h4>{funnel_html}
<details><summary>Final context preview ({len(final_chunks)} chunks)</summary>
{final_context_html or '<p>No final context.</p>'}</details>
<h4>Decision summary</h4><ul class="diagnostic-list">{decisions_html or '<li>No diagnostics.</li>'}</ul>
</article>"""


def write_standalone_html(
    path: str | Path,
    *,
    rows: Sequence[Mapping[str, Any]],
    responses: Sequence[Mapping[str, Any] | None],
    summary: Mapping[str, Any],
    metrics: Mapping[str, Any],
    title: str = "CIAL Knowledge OS — Phase 3 Hybrid Retrieval",
) -> Path:
    """Write one offline report with embedded styles, data, and charts."""

    question_sections: list[str] = []
    trace_sections: list[str] = []
    diagnostic_sections: list[str] = []
    token_sections: list[str] = []
    latency_sections: list[str] = []
    citation_sections: list[str] = []
    latency_bars: list[str] = []
    max_latency = max(
        (float(row.get("total_latency_seconds") or 0.0) for row in rows),
        default=0.0,
    )
    for index, row in enumerate(rows, start=1):
        response = responses[index - 1] if index <= len(responses) else None
        response = response or {}
        citations = response.get("citations")
        citation_items = (
            [_citation_html(item) for item in citations if isinstance(item, Mapping)]
            if isinstance(citations, Sequence)
            and not isinstance(citations, (str, bytes))
            else []
        )
        trace_value = response.get("question_trace")
        trace = trace_value if isinstance(trace_value, Mapping) else {}
        if trace:
            trace_sections.append(_question_trace_html(index, trace))
            decisions = [
                value
                for value in (trace.get("decision_summary") or [])
                if isinstance(value, Mapping)
            ]
            diagnostic_sections.append(
                f"<article><h3>Q{index}: {html.escape(str(row.get('question') or ''))}</h3><ul class=\"diagnostic-list\">"
                + "".join(
                    '<li class="diagnostic-card"><strong>'
                    + html.escape(str(item.get("signal") or ""))
                    + "</strong><span>"
                    + html.escape(str(item.get("recommendation") or ""))
                    + "</span></li>"
                    for item in decisions
                )
                + "</ul></article>"
            )
            trace_usage = trace.get("token_usage")
            trace_usage = (
                trace_usage if isinstance(trace_usage, Mapping) else {}
            )
            token_sections.append(
                f"""<article><h3>Q{index}: Token usage</h3><div class="grid">
{_metric_card("Budget", trace_usage.get("max_context_tokens", ""))}
{_metric_card("Used", trace_usage.get("context_tokens_used", 0))}
{_metric_card("Remaining", trace_usage.get("remaining_tokens", ""))}
{_metric_card("Utilization", f"{trace_usage.get('utilization_percent', 0)}%")}
{_metric_card("Chunks included", trace_usage.get("chunks_included", 0))}
{_metric_card("Chunks skipped", trace_usage.get("chunks_skipped", 0))}
</div></article>"""
            )
            trace_latency = trace.get("latency")
            trace_latency = (
                trace_latency if isinstance(trace_latency, Mapping) else {}
            )
            latency_sections.append(
                f"<article><h3>Q{index}: Latency breakdown</h3>"
                + _table_html(
                    ["Stage", "Seconds"],
                    [
                        [key.replace("_", " ").title(), value]
                        for key, value in trace_latency.items()
                        if value is not None
                    ],
                )
                + "</article>"
            )
        citation_sections.append(
            f'<article><h3>Q{index}: {html.escape(str(row.get("question") or ""))}</h3>'
            f'<ul class="citation-list">{"".join(citation_items) or "<li class=\"citation-card\">No citations</li>"}</ul></article>'
        )
        retrieved = response.get("context_stages")
        retrieved = (
            retrieved.get("compressed", [])
            if isinstance(retrieved, Mapping)
            else []
        )
        context_blocks = "".join(
            "<details><summary>"
            + html.escape(
                f"{item.get('source') or 'Unknown'} — "
                f"{item.get('chunk_id') or 'chunk unknown'}"
            )
            + "</summary><pre>"
            + html.escape(str(item.get("text") or ""))
            + "</pre></details>"
            for item in retrieved
            if isinstance(item, Mapping)
        )
        usage = response.get("token_usage")
        usage = usage if isinstance(usage, Mapping) else {}
        answer_text = _answer_without_reference_tail(
            str(row.get("answer") or ""),
            citations
            if isinstance(citations, Sequence)
            and not isinstance(citations, (str, bytes))
            else [],
        )
        if usage.get("budget_type") == "tokens":
            usage_label = (
                f"{usage.get('used', 0)} / {usage.get('budget', 0)} tokens "
                f"({usage.get('encoding_name', '')})"
            )
        else:
            usage_label = (
                f"{usage.get('context_tokens', usage.get('used', 0))} tokens "
                f"({usage.get('encoding_name', '')}); "
                f"{usage.get('characters_used', 0)} / "
                f"{usage.get('character_budget', 0)} legacy characters"
            )
        question_sections.append(
            f"""<article>
<h3>{index}. {html.escape(str(row.get("question") or ""))}</h3>
<div class="status">{html.escape(str(row.get("answer_status") or row.get("status") or ""))}</div>
<h4>Answer</h4><div class="answer-content">{_render_answer_markdown(answer_text)}</div>
<h4>Citations</h4><ul class="citation-list">{''.join(citation_items) or '<li class="citation-card">No citations</li>'}</ul>
<h4>Retrieved Context</h4>{context_blocks or '<p>No retained context.</p>'}
<div class="grid">
{_metric_card("Retrieved chunks", row.get("retrieved_chunks", 0))}
{_metric_card("Final sections", row.get("final_context_sections", 0))}
{_metric_card("Context usage", usage_label)}
{_metric_card("Latency (s)", row.get("total_latency_seconds", 0))}
</div>
</article>"""
        )
        latency = float(row.get("total_latency_seconds") or 0.0)
        width = 0 if max_latency == 0 else round(100 * latency / max_latency, 2)
        latency_bars.append(
            f'<div class="bar-row"><span>Q{index}</span><i style="width:{width}%"></i>'
            f"<b>{latency:.3f}s</b></div>"
        )

    card_values = [
            ("Questions", summary.get("question_count", len(rows))),
            ("Successful", summary.get("successful_questions", 0)),
            ("Answered", summary.get("answered_questions", 0)),
            ("Safe failures", summary.get("insufficient_evidence_questions", 0)),
            ("Average latency", summary.get("average_latency_seconds", 0)),
            ("Retrieval mode", summary.get("retrieval_mode", "")),
    ]
    card_values.extend(
        (key.replace("_", " ").title(), summary[key])
        for key in ("run_type", "run_label")
        if key in summary
    )
    cards = "".join(
        _metric_card(label, value)
        for label, value in card_values
    )
    embedded_data = json.dumps(
        {"summary": summary, "metrics": metrics},
        ensure_ascii=False,
        default=str,
    ).replace("</", "<\\/")
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{html.escape(title)}</title>
<style>
:root{{--navy:#12344d;--blue:#1f6f8b;--ice:#eef6f8;--ink:#18252d;--muted:#62737d}}
*{{box-sizing:border-box}}body{{margin:0;background:#f4f7f8;color:var(--ink);
font:15px/1.55 system-ui,-apple-system,Segoe UI,sans-serif}}
header{{background:linear-gradient(120deg,var(--navy),var(--blue));color:white;padding:42px max(5vw,24px)}}
main{{max-width:1200px;margin:auto;padding:28px}}section,article{{background:white;border:1px solid #dce5e8;
border-radius:12px;padding:22px;margin:18px 0;box-shadow:0 2px 8px #1232}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px}}
.metric{{background:var(--ice);border-left:4px solid var(--blue);padding:12px;border-radius:7px}}
.metric span{{display:block;color:var(--muted);font-size:12px;text-transform:uppercase}}.metric strong{{font-size:20px}}
pre{{white-space:pre-wrap;overflow-wrap:anywhere;background:#f7f9fa;padding:14px;border-radius:7px}}
.answer-content{{font-size:16px;line-height:1.7;color:var(--ink);background:#fbfdfe;
border-left:4px solid var(--blue);padding:16px 18px;border-radius:7px}}
.answer-content p{{margin:.35em 0 1em}}.answer-content h1,.answer-content h2,
.answer-content h3,.answer-content h4{{color:var(--navy);margin:1em 0 .45em}}
.answer-content ul,.answer-content ol{{padding-left:1.5rem;margin:.5em 0 1em}}
.answer-content li{{margin:.35em 0}}.answer-content code{{background:#e8f0f3;
padding:.12em .35em;border-radius:4px;font:0.92em ui-monospace,SFMono-Regular,Consolas,monospace}}
.citation-list{{list-style:none;padding:0;display:grid;gap:10px}}
.citation-card{{display:grid;grid-template-columns:auto 1fr;gap:12px;align-items:start;
background:var(--ice);border:1px solid #cddde2;border-radius:8px;padding:12px 14px}}
.citation-reference{{font-weight:700;color:var(--blue)}}.source-meta{{color:var(--muted);
font-size:13px;margin-top:3px}}.citation-action{{margin-top:5px}}.no-link{{color:var(--muted)}}
.table-wrap{{overflow-x:auto}}table{{width:100%;border-collapse:collapse;margin:10px 0 18px}}
th,td{{border:1px solid #dce5e8;padding:8px 10px;text-align:left;vertical-align:top}}
th{{background:var(--navy);color:white}}.trace-article details{{margin:12px 0}}
.trace-article summary{{font-weight:700;cursor:pointer;color:var(--navy)}}
.diagnostic-list{{list-style:none;padding:0;display:grid;gap:8px}}
.diagnostic-card{{display:grid;grid-template-columns:minmax(130px,180px) 1fr;gap:12px;
background:#f7fafb;border-left:4px solid var(--blue);padding:10px 12px;border-radius:6px}}
.status{{display:inline-block;background:#dbeef2;color:#16495b;padding:4px 9px;border-radius:999px}}
.bar-row{{display:grid;grid-template-columns:36px 1fr 72px;gap:8px;align-items:center;margin:8px 0}}
.bar-row i{{display:block;min-width:2px;height:18px;background:var(--blue);border-radius:3px}}
details{{border:1px solid #dce5e8;border-radius:6px;padding:8px;margin:8px 0}}a{{color:#0c607a}}
</style></head><body>
<header><h1>{html.escape(title)}</h1><p>Standalone, offline, evidence-aware run report.</p></header>
<main>
<section><h2>Executive Summary</h2><div class="grid">{cards}</div></section>
<section><h2>Metrics</h2><pre>{html.escape(json.dumps(metrics, indent=2, ensure_ascii=False, default=str))}</pre></section>
<section><h2>Question Answers</h2>
{''.join(question_sections) or '<p>No questions were processed.</p>'}</section>
<section><h2>Per-Question Trace</h2>{''.join(trace_sections) or '<p>No detailed traces recorded.</p>'}</section>
<section><h2>Retrieval Diagnostics</h2>{''.join(diagnostic_sections) or '<p>No diagnostics recorded.</p>'}</section>
<section><h2>Token Usage</h2>{''.join(token_sections) or '<p>No token traces recorded.</p>'}</section>
<section><h2>Latency Breakdown</h2>{''.join(latency_sections)}
<h3>Question latency comparison</h3>{''.join(latency_bars) or '<p>No latency data.</p>'}</section>
<section><h2>Citation Evidence</h2>{''.join(citation_sections) or '<p>No citations recorded.</p>'}</section>
<script type="application/json" id="run-data">{embedded_data}</script>
</main></body></html>"""
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(document, encoding="utf-8")
    logger.info(
        "html_report_written",
        extra={"event": "report_generation", "path": str(target)},
    )
    return target
