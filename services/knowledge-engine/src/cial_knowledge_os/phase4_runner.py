"""Phase 4 run orchestration built on the Phase 3 artifact lifecycle."""

from __future__ import annotations

import csv
import json
import warnings
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from statistics import fmean
from time import perf_counter
from typing import Any

from .benchmark_loader import Benchmark
from .batch_qa import _resolve_questions
from .config import Phase4Config
from .execution import ExecutionManager
from .file_formats import scan_file_format_readiness
from .phase3_runner import Phase3Runner
from .phase4_checkpoint import Phase4CheckpointManager
from .phase4_reporting import write_phase4_figures, write_phase4_html
from .run_manager import RunManager, RunPaths


@dataclass(frozen=True, slots=True)
class Phase4RunResult:
    """Expose one completed Phase 4 bundle and its aggregate diagnostics.

    ``paths`` is the unchanged Phase 3-compatible run path contract; ``summary``
    and ``metrics`` include additive reranking, evidence, token, and mode fields.
    ``run_mode`` records whether the caller executed smoke, manual QA,
    benchmark, or export-only behavior.
    """

    paths: RunPaths
    summary: dict[str, Any]
    metrics: dict[str, Any]
    run_mode: str


class Phase4Runner(Phase3Runner):
    """Generate Phase 4 artifacts while reusing the Phase 3 RunManager.

    Inputs are a Phase 4-compatible pipeline/config and optional RunManager.
    :meth:`run` produces the established CSV/XLSX/HTML/JSON/log/context bundle,
    then enriches summaries, figures, and HTML with Phase 4 diagnostics.

    The class subclasses :class:`Phase3Runner` intentionally: artifact paths,
    collision handling, batch failure isolation, citations, and legacy columns
    remain one implementation. Phase 4 only appends fields and overwrites the
    report with a richer standalone view.
    """

    def __init__(
        self,
        *,
        pipeline: Any,
        config: Phase4Config | None = None,
        run_manager: RunManager | None = None,
        execution_manager: ExecutionManager | None = None,
    ) -> None:
        phase4_config = config or pipeline.config
        if not isinstance(phase4_config, Phase4Config):
            raise TypeError("Phase4Runner requires a Phase4Config.")
        super().__init__(
            pipeline=pipeline,
            config=phase4_config,
            run_manager=run_manager,
        )
        self.config = phase4_config
        self.execution_manager = (
            execution_manager
            or ExecutionManager.from_config(
                phase4_config,
                phase="Phase 4",
                run_mode=phase4_config.phase4_run_mode,
            )
        )
        self.pipeline.execution_manager = self.execution_manager

    def _apply_mode_limits(
        self,
        questions: Sequence[str] | None,
        *,
        run_mode: str,
    ) -> Sequence[str] | None:
        if questions is None:
            return None
        values = list(questions)
        if run_mode == "smoke":
            return values[: min(3, len(values))]
        if run_mode in {"manual_qa", "export_only"}:
            limit = self.config.max_inline_manual_questions
            if (
                limit is not None
                and len(values) > limit
                and not self.config.allow_large_run
            ):
                warnings.warn(
                    "Phase 4 manual input exceeds "
                    f"{limit} questions; only the first configured limit will run. "
                    "Set allow_large_run=True and use export_only mode for an "
                    "intentional large batch.",
                    stacklevel=3,
                )
                return values[:limit]
        return values

    @staticmethod
    def _read_rows(path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]

    @staticmethod
    def _phase4_metrics(
        rows: Sequence[Mapping[str, Any]],
        traces: Sequence[Mapping[str, Any]],
    ) -> dict[str, Any]:
        def numbers(key: str) -> list[float]:
            values = []
            for row in rows:
                try:
                    values.append(float(row.get(key) or 0.0))
                except (TypeError, ValueError):
                    values.append(0.0)
            return values

        strengths: Counter[str] = Counter()
        discard_reasons: Counter[str] = Counter()
        diagnostic_flags: Counter[str] = Counter()
        fallback_questions = 0
        weak_evidence_questions = 0
        extractive_fallback_count = 0
        fallback_blocked_count = 0
        for trace in traces:
            quality = trace.get("evidence_quality")
            quality = quality if isinstance(quality, Mapping) else {}
            summary = quality.get("summary")
            summary = summary if isinstance(summary, Mapping) else {}
            strengths.update(summary.get("strength_distribution") or {})
            token_usage = trace.get("token_usage")
            token_usage = (
                token_usage if isinstance(token_usage, Mapping) else {}
            )
            discard_reasons.update(
                token_usage.get("discard_reason_distribution") or {}
            )
            fallback_questions += bool(token_usage.get("fallback_used"))
            weak_evidence_questions += bool(token_usage.get("weak_evidence"))
            extractive_fallback_count += bool(
                token_usage.get("extractive_fallback_used")
            )
            fallback_blocked_count += bool(
                token_usage.get("fallback_blocked")
            )
            diagnostic_flags.update(
                str(item.get("signal") or "unspecified")
                for item in (trace.get("decision_summary") or [])
                if isinstance(item, Mapping)
            )
        reductions = numbers("token_reduction_percent")
        reranker_scores = numbers("average_reranker_score")
        statuses = Counter(
            str(row.get("answer_status") or "")
            .strip()
            .casefold()
            .replace(" ", "_")
            for row in rows
        )
        return {
            "average_token_reduction_percent": (
                round(fmean(reductions), 6) if reductions else 0.0
            ),
            "average_reranker_score": (
                round(fmean(reranker_scores), 6) if reranker_scores else 0.0
            ),
            "selected_chunk_count": int(sum(numbers("selected_chunk_count"))),
            "discarded_chunk_count": int(sum(numbers("discarded_chunk_count"))),
            "average_selected_chunk_count": (
                round(fmean(numbers("selected_chunk_count")), 6) if rows else 0.0
            ),
            "average_discarded_chunk_count": (
                round(fmean(numbers("discarded_chunk_count")), 6) if rows else 0.0
            ),
            "average_reranker_latency_seconds": (
                round(fmean(numbers("reranker_latency_seconds")), 6)
                if rows
                else 0.0
            ),
            "average_selected_evidence_tokens": (
                round(fmean(numbers("selected_evidence_tokens")), 6)
                if rows
                else 0.0
            ),
            "average_citation_count": (
                round(fmean(numbers("citation_count")), 6) if rows else 0.0
            ),
            "fallback_question_count": fallback_questions,
            "weak_evidence_question_count": weak_evidence_questions,
            "unsupported_query_count": statuses["unsupported_query"],
            "insufficient_evidence_count": statuses[
                "insufficient_evidence"
            ],
            "extractive_fallback_count": extractive_fallback_count,
            "fallback_blocked_count": fallback_blocked_count,
            "discard_reason_distribution": dict(
                sorted(discard_reasons.items())
            ),
            "diagnostic_flag_counts": dict(
                sorted(diagnostic_flags.items())
            ),
            "evidence_strength_distribution": {
                name: strengths.get(name, 0)
                for name in ("strong", "medium", "weak")
            },
        }

    @staticmethod
    def _write_dict_csv(
        path: Path,
        rows: Sequence[Mapping[str, Any]],
        columns: Sequence[str],
    ) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(columns))
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def _file_format_summary_rows(readiness: Mapping[str, Any]) -> list[dict[str, Any]]:
        keys = (
            "total_files",
            "processable_files",
            "ocr_files",
            "recognized_future_files",
            "unsupported_files",
        )
        return [{"metric": key, "value": readiness.get(key, 0)} for key in keys]

    @staticmethod
    def _extension_rows(readiness: Mapping[str, Any]) -> list[dict[str, Any]]:
        rows = []
        for item in readiness.get("extensions") or []:
            if not isinstance(item, Mapping):
                continue
            rows.append(
                {
                    "extension": item.get("extension", ""),
                    "count": item.get("count", 0),
                    "category": item.get("category", ""),
                    "format_label": item.get("format_label", ""),
                    "support_status": item.get("support_status", ""),
                    "ingestion_enabled": item.get("ingestion_enabled", False),
                    "requires_ocr": item.get("requires_ocr", False),
                }
            )
        return rows

    @staticmethod
    def _skipped_rows(readiness: Mapping[str, Any]) -> list[dict[str, Any]]:
        rows = []
        for item in readiness.get("skipped_files") or []:
            if not isinstance(item, Mapping):
                continue
            rows.append(
                {
                    "filename": item.get("filename", ""),
                    "path": item.get("path", ""),
                    "extension": item.get("extension", ""),
                    "category": item.get("category", ""),
                    "support_status": item.get("support_status", ""),
                    "reason": item.get("reason", ""),
                    "action": item.get("action", ""),
                }
            )
        return rows

    def _write_file_format_exports(
        self,
        run_root: Path,
        workbook_path: Path,
        readiness: Mapping[str, Any],
    ) -> None:
        summary_rows = self._file_format_summary_rows(readiness)
        extension_rows = self._extension_rows(readiness)
        skipped_rows = self._skipped_rows(readiness)
        self._write_dict_csv(
            run_root / "file_format_summary.csv",
            summary_rows,
            ("metric", "value"),
        )
        self._write_dict_csv(
            run_root / "file_extension_distribution.csv",
            extension_rows,
            (
                "extension",
                "count",
                "category",
                "format_label",
                "support_status",
                "ingestion_enabled",
                "requires_ocr",
            ),
        )
        self._write_dict_csv(
            run_root / "skipped_files.csv",
            skipped_rows,
            (
                "filename",
                "path",
                "extension",
                "category",
                "support_status",
                "reason",
                "action",
            ),
        )
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return
        workbook = load_workbook(workbook_path)
        sheet_specs = (
            ("file_format_summary", summary_rows),
            ("file_extension_distribution", extension_rows),
            ("skipped_files", skipped_rows),
        )
        for sheet_name, rows in sheet_specs:
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            sheet = workbook.create_sheet(sheet_name)
            columns = list(rows[0].keys()) if rows else ["message"]
            header_fill = PatternFill("solid", fgColor="1F4E78")
            for column_index, name in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=column_index, value=name)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
            if rows:
                for row_index, row in enumerate(rows, start=2):
                    for column_index, name in enumerate(columns, start=1):
                        sheet.cell(row=row_index, column=column_index, value=row.get(name, ""))
            else:
                sheet.cell(row=2, column=1, value="No records.")
            sheet.freeze_panes = "A2"
        workbook.save(workbook_path)

    def run(
        self,
        *,
        questions: list[str] | tuple[str, ...] | None = None,
        questions_path: str | Path | None = None,
        benchmark: Benchmark | None = None,
        top_k: int | None = None,
        run_metadata: Mapping[str, Any] | None = None,
        run_mode: str | None = None,
        resume_run: str | Path | None = None,
    ) -> Phase4RunResult:
        """Execute Phase 4 and write the complete compatible artifact bundle.

        Inputs match :class:`Phase3Runner` plus ``run_mode`` and optional
        ``resume_run``. Smoke mode caps an in-memory list at three questions.
        Manual/export-only runs are unlimited by default; setting
        ``max_inline_manual_questions`` to a positive integer restores an
        explicit safety cap unless ``allow_large_run`` is enabled. Benchmark
        mode is never silently truncated. Each attempt is checkpointed before
        final CSV/XLSX/HTML/JSON/context/figure exports. Resume validates indexed
        question hashes, skips successful occurrences, retries failed ones, and
        rebuilds the standard artifact bundle in the original run folder.
        """

        effective_mode = run_mode or self.config.phase4_run_mode
        allowed = {"smoke", "manual_qa", "benchmark", "export_only"}
        if effective_mode not in allowed:
            raise ValueError(
                "run_mode must be smoke, manual_qa, benchmark, or export_only."
            )
        mode_questions: Sequence[str] | None = questions
        if (
            mode_questions is None
            and benchmark is not None
        ):
            mode_questions = [item.question for item in benchmark.questions]
        if mode_questions is None and questions_path is not None:
            mode_questions = _resolve_questions(
                questions=None,
                questions_path=questions_path,
                project_root=self.config.project_root,
            )
            questions_path = None
        limited_questions = self._apply_mode_limits(
            mode_questions,
            run_mode=effective_mode,
        )
        if limited_questions is None:
            raise ValueError("Phase 4 requires questions or a benchmark.")
        all_questions = list(limited_questions)
        self.execution_manager.run_mode = effective_mode
        if resume_run is not None:
            self.run_manager = RunManager.from_existing(
                self.config,
                resume_run,
            )
        paths = self.run_manager.create()
        checkpoint = Phase4CheckpointManager(paths.root)
        checkpoint.initialize(
            all_questions,
            config=self.config,
            resume=resume_run is not None,
        )
        pending = checkpoint.pending()
        initial_rows, initial_responses = checkpoint.completed_records()
        print(f"Skipped due to resume: {len(initial_rows)}")
        print(f"Remaining question count: {len(pending)}")
        print(f"Checkpoint path: {checkpoint.checkpoint_json}")
        self.execution_manager.start_run(
            total_questions=len(pending),
            resumed=resume_run is not None,
        )

        def checkpoint_question(
            position: int,
            row: dict[str, Any],
            response: Mapping[str, Any] | None,
        ) -> None:
            checkpoint.record(pending[position - 1], row, response)
            self.execution_manager.write_checkpoint_event(
                checkpoint.checkpoint_json,
                question_index=position,
            )

        metadata = dict(run_metadata or {})
        metadata.setdefault("run_mode", effective_mode)
        indexing_summary = getattr(self.pipeline, "indexing_summary", None)
        if isinstance(indexing_summary, Mapping) and indexing_summary:
            metadata.setdefault("indexing_summary", dict(indexing_summary))
        try:
            phase3_result = super().run(
                questions=[identity.question for identity in pending],
                questions_path=None,
                benchmark=benchmark,
                top_k=top_k,
                run_metadata=metadata,
                initial_rows=initial_rows,
                initial_responses=initial_responses,
                on_question_complete=checkpoint_question,
            )
        except Exception as exc:
            self.execution_manager.fail_run(exc)
            raise
        started = perf_counter()
        rows = self._read_rows(phase3_result.paths.results_csv)
        traces_value = json.loads(
            phase3_result.paths.retrieval_json.read_text(encoding="utf-8")
        )
        traces = [
            dict(trace)
            for trace in traces_value
            if isinstance(trace, Mapping)
        ]
        file_format_readiness = scan_file_format_readiness(
            self.config.knowledge_root
        )
        ocr_summary = getattr(self.pipeline, "ocr_summary", None)
        if not isinstance(ocr_summary, Mapping):
            ocr_summary = {
                "total_ocr_files_processed": 0,
                "ocr_success_count": 0,
                "ocr_failure_count": 0,
                "ocr_success_rate": 0.0,
                "average_ocr_processing_time_ms": 0.0,
                "total_extracted_characters": 0,
                "total_extracted_words": 0,
                "ocr_engine_used": self.config.ocr_engine,
                "failures": [],
            }
        additions = self._phase4_metrics(rows, traces)
        summary = dict(phase3_result.summary) | {
            "phase": "Phase 4",
            "qualification_status": "implemented_not_benchmark_qualified",
            "run_mode": effective_mode,
            "average_token_reduction_percent": additions[
                "average_token_reduction_percent"
            ],
            "file_format_readiness": file_format_readiness,
            "ocr_summary": dict(ocr_summary),
        }
        metrics = dict(phase3_result.metrics) | additions | {
            "phase": "Phase 4",
            "run_mode": effective_mode,
            "file_format_readiness": file_format_readiness,
            "ocr_summary": dict(ocr_summary),
        }
        self._write_file_format_exports(
            phase3_result.paths.root,
            phase3_result.paths.results_xlsx,
            file_format_readiness,
        )
        write_phase4_figures(phase3_result.paths.figures, traces)
        elapsed = perf_counter() - started
        for trace in traces:
            latency = trace.get("latency")
            latency = dict(latency) if isinstance(latency, Mapping) else {}
            latency["artifact_export_seconds"] = round(
                float(latency.get("artifact_export_seconds") or 0.0) + elapsed,
                6,
            )
            trace["latency"] = latency
        self.run_manager.write_json(phase3_result.paths.summary_json, summary)
        self.run_manager.write_json(phase3_result.paths.metrics_json, metrics)
        self.run_manager.write_json(phase3_result.paths.retrieval_json, traces)
        write_phase4_html(
            phase3_result.paths.report_html,
            rows=rows,
            traces=traces,
            summary=summary,
            metrics=metrics,
        )
        checkpoint.finalize(
            {
                "results_csv": phase3_result.paths.results_csv,
                "results_xlsx": phase3_result.paths.results_xlsx,
                "report_html": phase3_result.paths.report_html,
                "config_json": phase3_result.paths.config_json,
                "summary_json": phase3_result.paths.summary_json,
                "metrics_json": phase3_result.paths.metrics_json,
                "retrieval_json": phase3_result.paths.retrieval_json,
                "logs": phase3_result.paths.logs,
                "context": phase3_result.paths.context,
                "figures": phase3_result.paths.figures,
            }
        )
        self.execution_manager.emit(
            "batch_completed",
            status="completed",
            payload={
                "artifact_root": str(phase3_result.paths.root),
                "question_count": len(rows),
            },
            source="phase4_runner",
        )
        self.execution_manager.complete_run(
            artifact_root=str(phase3_result.paths.root),
        )
        return Phase4RunResult(
            paths=phase3_result.paths,
            summary=summary,
            metrics=metrics,
            run_mode=effective_mode,
        )
